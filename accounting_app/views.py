
from django.db import transaction
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, FileResponse, HttpResponse
from .models import JournalEntry, JournalLine, Account, AccountingPeriod
from .models import Customer, Supplier, SalesInvoice, PurchaseInvoice, Payment
from .forms import CustomerForm, SupplierForm, SalesInvoiceForm, PurchaseInvoiceForm, SalesItemFormSet, PurchaseItemFormSet, PaymentForm
from django.core.exceptions import ValidationError

from .forms import JournalEntryForm, JournalLineFormSet, AccountForm


from django.views.decorators.http import require_POST


# ==================================
# تقارير: مستندات غير مرحّلة
# ==================================
@login_required
def unposted_documents(request):
    """تقرير موحّد للمستندات غير المرحّلة (بيع/شراء/سندات)."""
    date_from = _parse_date(request.GET.get("date_from"))
    date_to = _parse_date(request.GET.get("date_to"))

    sales_qs = SalesInvoice.objects.select_related("customer").filter(journal_entry__isnull=True)
    purchase_qs = PurchaseInvoice.objects.select_related("supplier").filter(journal_entry__isnull=True)
    payments_qs = Payment.objects.select_related("customer", "supplier").filter(journal_entry__isnull=True)

    if date_from:
        sales_qs = sales_qs.filter(date__gte=date_from)
        purchase_qs = purchase_qs.filter(date__gte=date_from)
        payments_qs = payments_qs.filter(date__gte=date_from)
    if date_to:
        sales_qs = sales_qs.filter(date__lte=date_to)
        purchase_qs = purchase_qs.filter(date__lte=date_to)
        payments_qs = payments_qs.filter(date__lte=date_to)

    return render(request, "accounting_app/unposted_documents.html", {
        "date_from": request.GET.get("date_from", ""),
        "date_to": request.GET.get("date_to", ""),
        "sales_invoices": sales_qs.order_by("-date", "-id")[:2000],
        "purchase_invoices": purchase_qs.order_by("-date", "-id")[:2000],
        "payments": payments_qs.order_by("-date", "-id")[:2000],
    })


# ==================================
# عكس مستندات (فواتير/سندات)
# ==================================
@login_required
@require_POST
def reverse_sales_invoice(request, invoice_id: int):
    inv = get_object_or_404(SalesInvoice.objects.select_related("journal_entry__period"), id=invoice_id)
    if not inv.journal_entry_id:
        messages.error(request, "لا يوجد قيد مرتبط بهذه الفاتورة (غير مرحّلة).")
        return redirect("account:sales_invoices")

    entry = inv.journal_entry
    if entry.period and entry.period.is_closed:
        messages.error(request, "لا يمكن عمل قيد عكسي داخل فترة محاسبية مقفلة.")
        return redirect("account:sales_invoices")

    # إعادة استخدام منطق عكس القيد
    return reverse_journal_entry(request, entry.id)


@login_required
@require_POST
def reverse_purchase_invoice(request, invoice_id: int):
    inv = get_object_or_404(PurchaseInvoice.objects.select_related("journal_entry__period"), id=invoice_id)
    if not inv.journal_entry_id:
        messages.error(request, "لا يوجد قيد مرتبط بهذه الفاتورة (غير مرحّلة).")
        return redirect("account:purchase_invoices")

    entry = inv.journal_entry
    if entry.period and entry.period.is_closed:
        messages.error(request, "لا يمكن عمل قيد عكسي داخل فترة محاسبية مقفلة.")
        return redirect("account:purchase_invoices")

    return reverse_journal_entry(request, entry.id)


@login_required
@require_POST
def reverse_payment(request, payment_id: int):
    p = get_object_or_404(Payment.objects.select_related("journal_entry__period"), id=payment_id)
    if not p.journal_entry_id:
        messages.error(request, "لا يوجد قيد مرتبط بهذا السند (غير مرحّل).")
        return redirect("account:payments")

    entry = p.journal_entry
    if entry.period and entry.period.is_closed:
        messages.error(request, "لا يمكن عمل قيد عكسي داخل فترة محاسبية مقفلة.")
        return redirect("account:payments")

    return reverse_journal_entry(request, entry.id)
from django.db.models import Sum, Q
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.exceptions import ValidationError
from django.db import transaction

from .models import OpeningBalance
from .forms import OpeningBalanceFormSet


from .forms import JournalEntryForm, AccountForm

import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from bidi.algorithm import get_display
import arabic_reshaper
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from datetime import datetime
import openpyxl

from .models import JournalEntry, JournalLine, Account



from django.contrib.admin.views.decorators import staff_member_required
from django.views.decorators.http import require_POST


import io
import os
from datetime import datetime

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.http import FileResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from .models import Payment

@login_required
@staff_member_required
def periods_list(request):
    periods = AccountingPeriod.objects.all().order_by("-start_date")
    return render(request, "accounting_app/periods_list.html", {"periods": periods})


@login_required
@staff_member_required
@require_POST
def period_toggle_close(request, period_id):
    period = get_object_or_404(AccountingPeriod, id=period_id)

    action = request.POST.get("action")  # "close" or "open"
    if action == "close":
        period.is_closed = True
        period.save(update_fields=["is_closed"])
        messages.success(request, f"تم إقفال الفترة: {period.name}")
    elif action == "open":
        period.is_closed = False
        period.save(update_fields=["is_closed"])
        messages.success(request, f"تم فتح الفترة: {period.name}")
    else:
        messages.error(request, "طلب غير صحيح")

    return redirect("account:periods_list")



# ===============================
# Login / Logout / Dashboard
# ===============================
def login_view(request):
    if request.user.is_authenticated:
        return redirect('ui:dashboard')

    if request.method == 'POST':
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect('ui:dashboard')

        else:
            messages.error(request, "اسم المستخدم أو كلمة المرور غير صحيحة")
    return render(request, "login.html")



@login_required
def logout_view(request):
    logout(request)
    return redirect("account:login")



@login_required
def dashboard_view(request):
    return render(request, "dashboard.html")




# ===============================
# الصفحة الرئيسية للحسابات
# ===============================
@login_required
def accounting_home(request):
    operations = [
        {"name": "القيود اليومية", "url": "journal_entries"},
        {"name": "دفتر الأستاذ (حساب)", "url": "ledger_account"},   # (مربوط على general_ledger عندك)
        {"name": "ميزان المراجعة", "url": "trial_balance"},
        {"name": "قائمة الدخل", "url": "income_statement"},
        {"name": "الميزانية", "url": "balance_sheet"},
        {"name": "فواتير المبيعات", "url": "sales_invoices"},
        {"name": "فواتير المشتريات", "url": "purchase_invoices"},
        {"name": "حسابات العملاء", "url": "customer_accounts"},
        {"name": "حسابات الموردين", "url": "supplier_accounts"},
        {"name": "شجرة الحسابات", "url": "chart_of_accounts"},
        {"name": "إدارة النقدية", "url": "cash_management"},
        {"name": "سندات القبض والصرف", "url": "payments"},

        {"name": "مستندات غير مرحّلة", "url": "unposted_documents"},

        {"name": "الأرصدة الافتتاحية", "url": "opening_balances"},
        {"name": "إدارة الفترات المحاسبية", "url": "periods_list"},
    ]
    return render(request, "accounting_app/accounting_home.html", {"operations": operations})



# ===============================
# القيود اليومية
# ===============================
from django.core.paginator import Paginator

@login_required
def get_account_name(request):
    account_id = request.GET.get("account_id")
    if not account_id:
        return JsonResponse({"name": ""})

    acc = Account.objects.filter(id=account_id).values("name").first()
    return JsonResponse({"name": acc["name"] if acc else ""})


@login_required
def journal_entries(request):
    if request.method == "POST":
        entry_form = JournalEntryForm(request.POST)
        line_formset = JournalLineFormSet(request.POST)

        if entry_form.is_valid() and line_formset.is_valid():
            entry = entry_form.save(commit=False)

            if entry.period_id and entry.period.is_closed:
                messages.error(request, "لا يمكن إضافة قيد في فترة محاسبية مُقفلة.")
                return redirect("account:journal_entries")

            entry.created_by = request.user

            try:
                entry.save()
            except ValidationError as e:
                messages.error(request, str(e))
                return redirect("account:journal_entries")

            lines = line_formset.save(commit=False)

            total_debit = 0.0
            total_credit = 0.0

            try:
                for line in lines:
                    line.entry = entry
                    total_debit += float(line.debit or 0)
                    total_credit += float(line.credit or 0)
                    line.save()

                for obj in line_formset.deleted_objects:
                    obj.delete()

                if round(total_debit, 2) != round(total_credit, 2):
                    entry.delete()
                    messages.error(request, "القيد غير متوازن (المدين لا يساوي الدائن).")
                else:
                    messages.success(request, "تمت إضافة القيد بنجاح.")
                    return redirect("account:journal_entries")

            except ValidationError as e:
                entry.delete()
                messages.error(request, str(e))
                return redirect("account:journal_entries")

    else:
        entry_form = JournalEntryForm()
        line_formset = JournalLineFormSet()

    # =========================
    # ✅ بحث القيود (لا نعرض شيء افتراضيًا)
    # =========================
    q = (request.GET.get("q") or "").strip()
    date_from = (request.GET.get("date_from") or "").strip()
    date_to = (request.GET.get("date_to") or "").strip()

    entries = None
    page_obj = None

    if q or date_from or date_to:
        qs = (
            JournalEntry.objects
            .select_related("period")
            .prefetch_related("lines__account")
            .order_by("-id")
        )

        if q:
            qs = qs.filter(
                Q(serial_number__icontains=q) |
                Q(reference__icontains=q) |
                Q(description__icontains=q)
            )

        if date_from:
            qs = qs.filter(date__gte=date_from)

        if date_to:
            qs = qs.filter(date__lte=date_to)

        paginator = Paginator(qs, 30)
        page_number = request.GET.get("page")
        page_obj = paginator.get_page(page_number)
        entries = page_obj

    context = {
        "entry_form": entry_form,
        "line_formset": line_formset,

        # ✅ مهم: افتراضيًا None (يعني لا نعرض شيء)
        "entries": entries,
        "page_obj": page_obj,

        # للحفاظ على قيم البحث داخل الفورم
        "q": q,
        "date_from": date_from,
        "date_to": date_to,
    }
    return render(request, "accounting_app/journal_entries.html", context)


# ===============================
# القيد العكسي
# ===============================
@login_required
@require_POST
@transaction.atomic
@login_required
@require_POST
def reverse_journal_entry(request, entry_id):
    entry = get_object_or_404(
        JournalEntry.objects.select_related("period").prefetch_related("lines__account"),
        id=entry_id
    )

    if JournalEntry.objects.filter(reversed_entry=entry).exists():
        messages.error(request, "لا يمكن عمل قيد عكسي لقيد عكسي.")
        return redirect("account:journal_entries")

    if entry.is_reversed or entry.reversed_entry_id:
        messages.info(request, "هذا القيد تم عكسه مسبقًا.")
        return redirect("account:journal_entries")

    if entry.period and entry.period.is_closed:
        messages.error(request, "لا يمكن عمل قيد عكسي داخل فترة محاسبية مقفلة.")
        return redirect("account:journal_entries")

    ref = entry.serial_number or str(entry.id)
    rev = JournalEntry.objects.create(
        period=entry.period,
        date=entry.date,
        reference=f"REV-{ref}",
        description=f"قيد عكسي للقيد: {ref} | {entry.description}",
        created_by=request.user,
    )

    for line in entry.lines.all():
        JournalLine.objects.create(
            entry=rev,
            account=line.account,
            debit=line.credit or 0,
            credit=line.debit or 0,
            note=f"عكس: {line.note}" if line.note else f"عكس قيد {ref}",
        )

    entry.is_reversed = True
    entry.reversed_entry = rev
    entry.save(update_fields=["is_reversed", "reversed_entry"])

    messages.success(request, f"تم إنشاء القيد العكسي للقيد {ref} بنجاح.")
    return redirect("account:journal_entries")


# ===============================
# تصدير القيود PDF
# ===============================
@login_required
@login_required
def export_journal_pdf(request):
    entries = (JournalEntry.objects.prefetch_related("lines__account").order_by("-id")[:2000] )  # حد أعلى للتصدير



    buffer = io.BytesIO()
    width, height = landscape(A4)
    pdf = canvas.Canvas(buffer, pagesize=landscape(A4))

    pdfmetrics.registerFont(TTFont('Arial', 'arial.ttf'))

    pdf.setFont("Arial", 18)
    header_text = get_display(arabic_reshaper.reshape("مصنع المحبة للصناعات الغذائية"))
    pdf.drawCentredString(width / 2, height - 40, header_text)

    pdf.setFont("Arial", 14)
    date_text = get_display(
        arabic_reshaper.reshape(f"تاريخ الطباعة: {datetime.today().strftime('%Y-%m-%d')}")
    )
    pdf.drawCentredString(width / 2, height - 65, date_text)

    styles = getSampleStyleSheet()
    arabic_style = ParagraphStyle(
        'arabic',
        parent=styles['Normal'],
        fontName='Arial',
        fontSize=12,
        alignment=1,
        leading=14
    )

    headers = ["الرقم المسلسل", "تاريخ المستند", "رقم الحساب", "اسم الحساب", "البيان", "مدين", "دائن"]
    data = [[Paragraph(get_display(arabic_reshaper.reshape(h)), arabic_style) for h in headers]]

    for entry in entries:
        for line in entry.lines.all():
            row = [
                Paragraph(str(entry.serial_number), arabic_style),
                Paragraph(entry.date.strftime("%Y-%m-%d"), arabic_style),
                Paragraph(str(line.account.code), arabic_style),
                Paragraph(get_display(arabic_reshaper.reshape(line.account.name)), arabic_style),
                Paragraph(
                    get_display(arabic_reshaper.reshape(line.note or entry.description)),
                    arabic_style
                ),
                Paragraph(str(line.debit), arabic_style),
                Paragraph(str(line.credit), arabic_style),
            ]
            data.append(row)

    col_widths = [
        width * 0.08,
        width * 0.12,
        width * 0.12,
        width * 0.20,
        width * 0.20,
        width * 0.14,
        width * 0.14,
    ]

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.whitesmoke),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOX', (0, 0), (-1, -1), 1, colors.grey),
    ]))

    table.wrapOn(pdf, width, height)
    table_height = table._height
    table.drawOn(pdf, (width - sum(col_widths)) / 2, height - 100 - table_height)

    pdf.setFont("Arial", 10)
    page_text = get_display(arabic_reshaper.reshape("صفحة 1"))
    pdf.drawCentredString(width / 2, 20, page_text)

    pdf.showPage()
    pdf.save()

    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename='journal_entries.pdf')


import io
import os
from datetime import datetime

from django.http import FileResponse
from django.shortcuts import get_object_or_404
from django.conf import settings
from django.utils import timezone
from django.contrib.auth.decorators import login_required

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

import arabic_reshaper
from bidi.algorithm import get_display

from .models import JournalEntry


def _ar(txt):
    return get_display(arabic_reshaper.reshape(str(txt or "")))


def _register_arabic_font():
    candidates = [
        os.path.join(settings.BASE_DIR, "accounting_app", "static", "fonts", "Amiri-Regular.ttf"),
        os.path.join(settings.BASE_DIR, "inventory", "static", "fonts", "Amiri-Regular.ttf"),
        os.path.join(settings.BASE_DIR, "static", "fonts", "Amiri-Regular.ttf"),
        r"C:\Windows\Fonts\arial.ttf",
    ]
    for p in candidates:
        if p and os.path.exists(p):
            try:
                pdfmetrics.registerFont(TTFont("AR_FONT", p))
                return "AR_FONT"
            except Exception:
                continue
    return "Helvetica"


@login_required
def export_single_journal_pdf(request, entry_id):
    entry = get_object_or_404(JournalEntry.objects.prefetch_related('lines', 'lines__account'), id=entry_id)

    font_name = _register_arabic_font()

    company_name = "شركة المحبة للصناعات الغذائية"
    company_address = "الأردن - عمّان - أبو علندا"
    now_str = timezone.localtime(timezone.now()).strftime("%Y-%m-%d %H:%M")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=24, rightMargin=24, topMargin=70, bottomMargin=45
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title_ar", parent=styles["Title"], fontName=font_name, fontSize=16, alignment=1)
    small_style = ParagraphStyle("small_ar", parent=styles["Normal"], fontName=font_name, fontSize=10, alignment=1)
    cell_style = ParagraphStyle("cell_ar", parent=styles["Normal"], fontName=font_name, fontSize=11, alignment=1)

    def header_footer(canvas, doc_):
        canvas.saveState()
        canvas.setFont(font_name, 12)
        canvas.drawCentredString(doc_.pagesize[0] / 2, doc_.pagesize[1] - 25, _ar(company_name))
        canvas.setFont(font_name, 10)
        canvas.drawCentredString(doc_.pagesize[0] / 2, doc_.pagesize[1] - 42, _ar(company_address))
        canvas.drawCentredString(doc_.pagesize[0] / 2, doc_.pagesize[1] - 58, _ar(f"تاريخ الطباعة: {now_str}"))
        canvas.setFont(font_name, 10)
        canvas.drawRightString(doc_.pagesize[0] - 24, 18, _ar(f"صفحة {doc_.page}"))
        canvas.restoreState()

    elements = []

    serial_txt = entry.serial_number if entry.serial_number else entry.id
    period_txt = entry.period.name if getattr(entry, "period", None) else "-"
    ref_txt = entry.reference or "-"

    elements.append(Paragraph(_ar("تقرير قيد يومي"), title_style))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(_ar(f"رقم القيد: {serial_txt} | التاريخ: {entry.date} | الفترة: {period_txt} | المرجع: {ref_txt}"), small_style))
    elements.append(Paragraph(_ar(f"البيان: {entry.description}"), small_style))
    elements.append(Spacer(1, 10))

    headers = ["#", "رمز الحساب", "اسم الحساب", "ملاحظة", "مدين", "دائن"]
    data = [[Paragraph(_ar(h), cell_style) for h in headers]]

    total_debit = 0
    total_credit = 0

    lines = list(entry.lines.all())
    for i, line in enumerate(lines, start=1):
        debit = float(line.debit or 0)
        credit = float(line.credit or 0)
        total_debit += debit
        total_credit += credit

        data.append([
            Paragraph(_ar(i), cell_style),
            Paragraph(_ar(getattr(line.account, "code", "")), cell_style),
            Paragraph(_ar(getattr(line.account, "name", "")), cell_style),
            Paragraph(_ar(line.note or ""), cell_style),
            Paragraph(_ar(f"{debit:.2f}" if debit else "-"), cell_style),
            Paragraph(_ar(f"{credit:.2f}" if credit else "-"), cell_style),
        ])

    # إجمالي
    data.append([
        Paragraph(_ar(""), cell_style),
        Paragraph(_ar(""), cell_style),
        Paragraph(_ar(""), cell_style),
        Paragraph(_ar("الإجمالي"), cell_style),
        Paragraph(_ar(f"{total_debit:.2f}"), cell_style),
        Paragraph(_ar(f"{total_credit:.2f}"), cell_style),
    ])

    # عرض الجدول بعرض الصفحة بشكل مرتب
    page_w = landscape(A4)[0]
    table_w = page_w - 24 - 24  # margins
    col_ratios = [0.06, 0.14, 0.22, 0.30, 0.14, 0.14]
    col_widths = [table_w * r for r in col_ratios]

    table = Table(data, colWidths=col_widths, repeatRows=1, hAlign="CENTER")
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 11),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E6E6E6")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.whitesmoke, colors.white]),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F1F1F1")),
    ]))

    elements.append(table)

    doc.build(elements, onFirstPage=header_footer, onLaterPages=header_footer)
    buffer.seek(0)

    filename = f"journal_entry_{serial_txt}.pdf"
    return FileResponse(buffer, as_attachment=True, filename=filename)


# ===============================
# PDF: فاتورة مبيعات / مشتريات
# ===============================
from reportlab.lib.pagesizes import A4, landscape

@login_required
def export_sales_invoice_pdf(request, invoice_id):
    inv = get_object_or_404(
        SalesInvoice.objects.select_related("customer").prefetch_related("items__product"),
        id=invoice_id
    )

    font_name = _register_arabic_font()

    company_name = "شركة المحبة للصناعات الغذائية"
    company_address = "الأردن - عمّان - أبو علندا"
    now_str = timezone.localtime(timezone.now()).strftime("%Y-%m-%d %H:%M")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=24, rightMargin=24, topMargin=70, bottomMargin=45
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title_ar", parent=styles["Title"], fontName=font_name, fontSize=18, alignment=1)
    small_style = ParagraphStyle("small_ar", parent=styles["Normal"], fontName=font_name, fontSize=11, alignment=2)
    cell_style = ParagraphStyle("cell_ar", parent=styles["Normal"], fontName=font_name, fontSize=11, alignment=1)

    def header_footer(canvas, doc_):
        canvas.saveState()
        canvas.setFont(font_name, 12)
        canvas.drawCentredString(doc_.pagesize[0] / 2, doc_.pagesize[1] - 25, _ar(company_name))
        canvas.setFont(font_name, 10)
        canvas.drawCentredString(doc_.pagesize[0] / 2, doc_.pagesize[1] - 42, _ar(company_address))
        canvas.drawCentredString(doc_.pagesize[0] / 2, doc_.pagesize[1] - 58, _ar(f"تاريخ الطباعة: {now_str}"))
        canvas.setFont(font_name, 10)
        canvas.drawRightString(doc_.pagesize[0] - 24, 18, _ar(f"صفحة {doc_.page}"))
        canvas.restoreState()

    elements = []
    elements.append(Paragraph(_ar("فاتورة مبيعات"), title_style))
    elements.append(Spacer(1, 8))

    elements.append(Paragraph(_ar(
        f"رقم الفاتورة: {inv.invoice_number or inv.id} | التاريخ: {inv.date} | العميل: {inv.customer.name}"
    ), small_style))
    elements.append(Spacer(1, 10))

    headers = ["#", "الصنف", "الكمية", "السعر", "الإجمالي"]
    data = [[Paragraph(_ar(h), cell_style) for h in headers]]

    total = 0.0
    items = list(inv.items.all())
    for i, it in enumerate(items, start=1):
        line_total = float(it.qty) * float(it.price)
        total += line_total
        data.append([
            Paragraph(_ar(i), cell_style),
            Paragraph(_ar(it.product.name), cell_style),
            Paragraph(_ar(it.qty), cell_style),
            Paragraph(_ar(f"{float(it.price):.2f}"), cell_style),
            Paragraph(_ar(f"{line_total:.2f}"), cell_style),
        ])

    # إجمالي
    data.append([
        Paragraph(_ar(""), cell_style),
        Paragraph(_ar(""), cell_style),
        Paragraph(_ar(""), cell_style),
        Paragraph(_ar("الإجمالي"), cell_style),
        Paragraph(_ar(f"{float(inv.total):.2f}"), cell_style),
    ])

    page_w = landscape(A4)[0]
    table_w = page_w - 24 - 24
    col_ratios = [0.08, 0.44, 0.16, 0.16, 0.16]
    col_widths = [table_w * r for r in col_ratios]

    table = Table(data, colWidths=col_widths, repeatRows=1, hAlign="CENTER")
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 11),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E6E6E6")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.whitesmoke, colors.white]),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F1F1F1")),
    ]))

    elements.append(table)
    doc.build(elements, onFirstPage=header_footer, onLaterPages=header_footer)
    buffer.seek(0)

    filename = f"sales_invoice_{inv.invoice_number or inv.id}.pdf"
    return FileResponse(buffer, as_attachment=True, filename=filename)


@login_required
def export_purchase_invoice_pdf(request, invoice_id):
    inv = get_object_or_404(
        PurchaseInvoice.objects.select_related("supplier").prefetch_related("items__product"),
        id=invoice_id
    )

    font_name = _register_arabic_font()

    company_name = "شركة المحبة للصناعات الغذائية"
    company_address = "الأردن - عمّان - أبو علندا"
    now_str = timezone.localtime(timezone.now()).strftime("%Y-%m-%d %H:%M")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=24, rightMargin=24, topMargin=70, bottomMargin=45
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title_ar", parent=styles["Title"], fontName=font_name, fontSize=18, alignment=1)
    small_style = ParagraphStyle("small_ar", parent=styles["Normal"], fontName=font_name, fontSize=11, alignment=2)
    cell_style = ParagraphStyle("cell_ar", parent=styles["Normal"], fontName=font_name, fontSize=11, alignment=1)

    def header_footer(canvas, doc_):
        canvas.saveState()
        canvas.setFont(font_name, 12)
        canvas.drawCentredString(doc_.pagesize[0] / 2, doc_.pagesize[1] - 25, _ar(company_name))
        canvas.setFont(font_name, 10)
        canvas.drawCentredString(doc_.pagesize[0] / 2, doc_.pagesize[1] - 42, _ar(company_address))
        canvas.drawCentredString(doc_.pagesize[0] / 2, doc_.pagesize[1] - 58, _ar(f"تاريخ الطباعة: {now_str}"))
        canvas.setFont(font_name, 10)
        canvas.drawRightString(doc_.pagesize[0] - 24, 18, _ar(f"صفحة {doc_.page}"))
        canvas.restoreState()

    elements = []
    elements.append(Paragraph(_ar("فاتورة مشتريات"), title_style))
    elements.append(Spacer(1, 8))

    elements.append(Paragraph(_ar(
        f"رقم الفاتورة: {inv.invoice_number or inv.id} | التاريخ: {inv.date} | المورد: {inv.supplier.name}"
    ), small_style))
    elements.append(Spacer(1, 10))

    headers = ["#", "الصنف", "الكمية", "السعر", "الإجمالي"]
    data = [[Paragraph(_ar(h), cell_style) for h in headers]]

    items = list(inv.items.all())
    for i, it in enumerate(items, start=1):
        line_total = float(it.qty) * float(it.price)
        data.append([
            Paragraph(_ar(i), cell_style),
            Paragraph(_ar(it.product.name), cell_style),
            Paragraph(_ar(it.qty), cell_style),
            Paragraph(_ar(f"{float(it.price):.2f}"), cell_style),
            Paragraph(_ar(f"{line_total:.2f}"), cell_style),
        ])

    # إجمالي
    data.append([
        Paragraph(_ar(""), cell_style),
        Paragraph(_ar(""), cell_style),
        Paragraph(_ar(""), cell_style),
        Paragraph(_ar("الإجمالي"), cell_style),
        Paragraph(_ar(f"{float(inv.total):.2f}"), cell_style),
    ])

    page_w = landscape(A4)[0]
    table_w = page_w - 24 - 24
    col_ratios = [0.08, 0.44, 0.16, 0.16, 0.16]
    col_widths = [table_w * r for r in col_ratios]

    table = Table(data, colWidths=col_widths, repeatRows=1, hAlign="CENTER")
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 11),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E6E6E6")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.whitesmoke, colors.white]),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F1F1F1")),
    ]))

    elements.append(table)
    doc.build(elements, onFirstPage=header_footer, onLaterPages=header_footer)
    buffer.seek(0)

    filename = f"purchase_invoice_{inv.invoice_number or inv.id}.pdf"
    return FileResponse(buffer, as_attachment=True, filename=filename)

# ===============================
# تصدير القيود Excel
# ===============================
@login_required
@login_required
def export_journal_excel(request):
    entries = (JournalEntry.objects.prefetch_related("lines__account").order_by("-id")[:20000])


    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Journal Entries"

    headers = ["الرقم المسلسل", "تاريخ المستند", "رقم الحساب", "اسم الحساب", "البيان", "مدين", "دائن"]
    sheet.append(headers)

    for entry in entries:
        for line in entry.lines.all():
            sheet.append([
                entry.serial_number,
                entry.date.strftime("%Y-%m-%d"),
                line.account.code,
                line.account.name,
                line.note or entry.description,
                line.debit,
                line.credit,
            ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=journal_entries.xlsx'
    workbook.save(response)
    return response


# ===============================
# Auto-fill لاسم الحساب
# ===============================
@login_required
def get_account_name(request):
    account_id = request.GET.get('account_id')
    try:
        account = Account.objects.get(id=account_id)
        return JsonResponse({'name': account.name})
    except Account.DoesNotExist:
        return JsonResponse({'name': ''})


# ===============================
# شجرة الحسابات
# ===============================
@login_required
def chart_of_accounts(request):
    accounts = Account.objects.all().order_by("code")
    return render(request, "accounting_app/chart_of_accounts.html", {"accounts": accounts})


# ===============================
# دفتر الأستاذ والفواتير والتقارير
# ===============================

@login_required
def general_ledger(request):
    period_id = (request.GET.get("period") or "").strip()
    account_id = (request.GET.get("account") or "").strip()

    periods = AccountingPeriod.objects.all().order_by("-start_date")
    accounts = Account.objects.all().order_by("code")

    selected_period = None
    if period_id:
        selected_period = AccountingPeriod.objects.filter(id=period_id).first()

    # ✅ إذا ما اختار حساب: هذا "دفتر الأستاذ العام" (قائمة قيود/سطور حسب الفترة)
    if not account_id:
        lines = JournalLine.objects.select_related("entry", "account")

        if selected_period:
            lines = lines.filter(
                entry__date__gte=selected_period.start_date,
                entry__date__lte=selected_period.end_date
            )

        lines = lines.order_by("entry__date", "entry__id", "id")[:2000]  # حدّ للحماية

        rows = []
        total_debit = 0.0
        total_credit = 0.0

        for line in lines:
            debit = float(line.debit or 0)
            credit = float(line.credit or 0)
            total_debit += debit
            total_credit += credit

            entry = line.entry
            rows.append({
                "mode": "gl",
                "date": entry.date,
                "entry_serial": entry.serial_number or entry.id,
                "entry_id": entry.id,
                "reference": entry.reference or "-",
                "description": entry.description,
                "account_code": line.account.code,
                "account_name": line.account.name,
                "note": line.note or "",
                "debit": debit,
                "credit": credit,
            })

        return render(request, "accounting_app/general_ledger.html", {
            "mode": "gl",
            "periods": periods,
            "accounts": accounts,
            "selected_period": str(period_id),
            "selected_account": "",
            "rows": rows,
            "total_debit": f"{total_debit:.2f}",
            "total_credit": f"{total_credit:.2f}",
            "hint": "اختاري حساب لو بدك دفتر أستاذ (حساب) مع رصيد جاري.",
            # حقول وضع الحساب نخليها صفر
            "opening_balance_text": "0.00",
            "opening_balance_class": "",
        })

    # ✅ إذا اختار حساب: هذا "دفتر أستاذ (حساب)" بنفس الصفحة + رصيد جاري
    base_lines = JournalLine.objects.select_related("entry", "account").filter(account_id=account_id)

    opening_balance = 0.0
    if selected_period:
        opening_sum = base_lines.filter(entry__date__lt=selected_period.start_date).aggregate(
            d=Sum("debit"), c=Sum("credit")
        )
        opening_balance = float(opening_sum["d"] or 0) - float(opening_sum["c"] or 0)

        lines = base_lines.filter(
            entry__date__gte=selected_period.start_date,
            entry__date__lte=selected_period.end_date,
        ).order_by("entry__date", "entry__id", "id")
    else:
        lines = base_lines.order_by("entry__date", "entry__id", "id")

    def fmt_balance(bal: float):
        if bal > 0:
            return f"{bal:.2f} مدين", "text-success fw-bold"
        if bal < 0:
            return f"{abs(bal):.2f} دائن", "text-danger fw-bold"
        return "0.00", ""

    opening_text, opening_class = fmt_balance(opening_balance)

    rows = []
    total_debit = 0.0
    total_credit = 0.0
    running_balance = opening_balance

    for line in lines:
        debit = float(line.debit or 0)
        credit = float(line.credit or 0)
        total_debit += debit
        total_credit += credit
        running_balance += (debit - credit)

        bal_text, bal_class = fmt_balance(running_balance)

        entry = line.entry
        rows.append({
            "mode": "account",
            "date": entry.date,
            "entry_serial": entry.serial_number or entry.id,
            "entry_id": entry.id,
            "reference": entry.reference or "-",
            "description": entry.description,
            "note": line.note or "",
            "debit": debit,
            "credit": credit,
            "balance_text": bal_text,
            "balance_class": bal_class,
        })

    return render(request, "accounting_app/general_ledger.html", {
        "mode": "account",
        "periods": periods,
        "accounts": accounts,
        "selected_period": str(period_id),
        "selected_account": str(account_id),
        "rows": rows,
        "opening_balance_text": opening_text,
        "opening_balance_class": opening_class,
        "total_debit": f"{total_debit:.2f}",
        "total_credit": f"{total_credit:.2f}",
        "hint": "",
    })

    # -------------------------------------------------------
    # Opening balance = مجموع (مدين - دائن) قبل بداية الفترة
    # -------------------------------------------------------
    opening_balance = 0

    base_lines = JournalLine.objects.select_related("entry", "account").filter(account_id=account_id)

    if selected_period:
        opening_qs = base_lines.filter(entry__date__lt=selected_period.start_date)
        opening_sum = opening_qs.aggregate(
            d=Sum("debit"),
            c=Sum("credit"),
        )
        opening_balance = float(opening_sum["d"] or 0) - float(opening_sum["c"] or 0)

        # خطوط الفترة
        lines = base_lines.filter(
            entry__date__gte=selected_period.start_date,
            entry__date__lte=selected_period.end_date,
        ).order_by("entry__date", "entry__id", "id")
    else:
        # بدون فترة: نعتبر الافتتاحي = 0 ونجيب كل الخطوط
        lines = base_lines.order_by("entry__date", "entry__id", "id")

    # -------------------------------------------------------
    # Build rows + running balance
    # -------------------------------------------------------
    rows = []
    total_debit = 0.0
    total_credit = 0.0
    running_balance = opening_balance

    def fmt_balance(bal: float):
        if bal > 0:
            return f"{bal:.2f} مدين", "text-success fw-bold"
        elif bal < 0:
            return f"{abs(bal):.2f} دائن", "text-danger fw-bold"
        return "0.00", ""

    opening_text, opening_class = fmt_balance(opening_balance)

    for line in lines:
        debit = float(line.debit or 0)
        credit = float(line.credit or 0)

        total_debit += debit
        total_credit += credit
        running_balance += (debit - credit)

        bal_text, bal_class = fmt_balance(running_balance)

        entry = line.entry
        rows.append({
            "date": entry.date,
            "entry_serial": entry.serial_number or entry.id,
            "entry_id": entry.id,
            "reference": entry.reference or "-",
            "description": entry.description,
            "note": line.note or "",
            "debit": debit,
            "credit": credit,
            "balance_text": bal_text,
            "balance_class": bal_class,
        })

    context = {
        "periods": periods,
        "accounts": accounts,
        "selected_period": str(period_id),
        "selected_account": str(account_id),
        "rows": rows,
        "opening_balance_value": opening_balance,
        "opening_balance_text": opening_text,
        "opening_balance_class": opening_class,
        "total_debit": f"{total_debit:.2f}",
        "total_credit": f"{total_credit:.2f}",
        "hint": "",
    }
    return render(request, "accounting_app/general_ledger.html", context)




@login_required
def customer_accounts(request):
    if request.method == "POST":
        form = CustomerForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "تمت إضافة العميل بنجاح")
            return redirect("account:customer_accounts")
    else:
        form = CustomerForm()

    customers = Customer.objects.all().order_by("name")
    return render(request, "accounting_app/customer_accounts.html", {"form": form, "customers": customers})


@login_required
def supplier_accounts(request):
    if request.method == "POST":
        form = SupplierForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "تمت إضافة المورد بنجاح")
            return redirect("account:supplier_accounts")
    else:
        form = SupplierForm()

    suppliers = Supplier.objects.all().order_by("name")
    return render(request, "accounting_app/supplier_accounts.html", {"form": form, "suppliers": suppliers})


# ===============================
# فواتير المبيعات
# ===============================
@login_required

@login_required
def sales_invoices(request):
    if request.method == "POST":
        form = SalesInvoiceForm(request.POST)
        formset = SalesItemFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    inv = form.save()
                    formset.instance = inv
                    formset.save()
                    inv.recalc_total()
                    inv.post_to_journal(user=request.user)
                messages.success(request, f"تمت إضافة فاتورة مبيعات وترحيلها ({inv.invoice_number})")
                return redirect("account:sales_invoices")
            except ValidationError as e:
                messages.error(request, str(e))
        else:
            messages.error(request, "يرجى التأكد من بيانات الفاتورة وبنودها.")
    else:
        form = SalesInvoiceForm()
        formset = SalesItemFormSet()

    invoices = SalesInvoice.objects.select_related("customer").prefetch_related("items").order_by("-date", "-id")
    return render(
        request,
        "accounting_app/sales_invoices.html",
        {"form": form, "formset": formset, "invoices": invoices},
    )


from django.views.decorators.http import require_POST

@login_required
@require_POST
@transaction.atomic
@login_required
@require_POST
@transaction.atomic
def post_sales_invoice(request, invoice_id):
    inv = get_object_or_404(SalesInvoice, id=invoice_id)

    # ✅ إذا كانت مرحّلة مسبقًا لا تحاولي ترحيلها مرة ثانية
    if inv.journal_entry_id:
        messages.info(request, "هذه الفاتورة مرحّلة مسبقًا")
        return redirect("account:sales_invoices")

    try:
        inv.post_to_journal(user=request.user)
        messages.success(request, f"تم ترحيل الفاتورة {inv.invoice_number} بنجاح")
    except ValidationError as e:
        messages.error(request, str(e))

    return redirect("account:sales_invoices")


@login_required
@require_POST
@transaction.atomic
def post_payment(request, payment_id: int):
    p = get_object_or_404(Payment, id=payment_id)

    if p.journal_entry_id:
        messages.info(request, "هذا السند مرحّل مسبقًا.")
        return redirect("account:payments")

    try:
        p.post_to_journal(user=request.user)
        messages.success(request, f"تم ترحيل السند {p.voucher_number or p.id} بنجاح.")
    except ValidationError as e:
        messages.error(request, str(e))

    return redirect("account:payments")


# ===============================
# فواتير المشتريات
# ===============================
@login_required

@login_required
def purchase_invoices(request):
    if request.method == "POST":
        form = PurchaseInvoiceForm(request.POST)
        formset = PurchaseItemFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    inv = form.save()
                    formset.instance = inv
                    formset.save()
                    inv.recalc_total()
                    inv.post_to_journal(user=request.user)
                messages.success(request, f"تمت إضافة فاتورة مشتريات وترحيلها ({inv.invoice_number})")
                return redirect("account:purchase_invoices")
            except ValidationError as e:
                messages.error(request, str(e))
        else:
            messages.error(request, "يرجى التأكد من بيانات الفاتورة وبنودها.")
    else:
        form = PurchaseInvoiceForm()
        formset = PurchaseItemFormSet()

    invoices = PurchaseInvoice.objects.select_related("supplier").prefetch_related("items").order_by("-date", "-id")
    return render(
        request,
        "accounting_app/purchase_invoices.html",
        {"form": form, "formset": formset, "invoices": invoices},
    )

from django.db.models import Sum
from .models import JournalLine, Account


from django.views.decorators.http import require_POST

@login_required
@require_POST
@transaction.atomic
def post_purchase_invoice(request, invoice_id):
    inv = get_object_or_404(PurchaseInvoice, id=invoice_id)

    # ✅ إذا كانت مرحّلة مسبقًا لا تحاولي ترحيلها مرة ثانية
    if inv.journal_entry_id:
        messages.info(request, "هذه الفاتورة مرحّلة مسبقًا")
        return redirect("account:purchase_invoices")

    try:
        inv.post_to_journal(user=request.user)
        messages.success(request, f"تم ترحيل فاتورة المشتريات {inv.invoice_number} بنجاح")
    except ValidationError as e:
        messages.error(request, str(e))

    return redirect("account:purchase_invoices")



@login_required
def income_statement(request):
    revenues = []
    expenses = []

    total_revenue = 0
    total_expense = 0

    accounts = Account.objects.all().order_by('code')

    for acc in accounts:
        sums = JournalLine.objects.filter(account=acc).aggregate(
            debit_sum=Sum('debit'),
            credit_sum=Sum('credit')
        )

        debit = sums['debit_sum'] or 0
        credit = sums['credit_sum'] or 0

        # إيرادات (غالبًا دائن)
        if acc.code.startswith('4'):
            amount = credit - debit
            if amount != 0:
                revenues.append({
                    'code': acc.code,
                    'name': acc.name,
                    'amount': amount
                })
                total_revenue += amount

        # مصاريف (غالبًا مدين)
        elif acc.code.startswith('5'):
            amount = debit - credit
            if amount != 0:
                expenses.append({
                    'code': acc.code,
                    'name': acc.name,
                    'amount': amount
                })
                total_expense += amount

    net_income = total_revenue - total_expense

    context = {
        'revenues': revenues,
        'expenses': expenses,
        'total_revenue': total_revenue,
        'total_expense': total_expense,
        'net_income': net_income,
    }
    return render(request, 'accounting_app/income_statement.html', context)


from django.db.models import Sum
from .models import JournalLine, Account


@login_required
def balance_sheet(request):
    assets = []
    liabilities = []
    equity = []

    total_assets = 0
    total_liabilities = 0
    total_equity = 0

    accounts = Account.objects.all().order_by('code')

    # ✅ حساب صافي الربح (حسابات 4 و 5)
    net_income = 0
    for acc in accounts:
        sums = JournalLine.objects.filter(account=acc).aggregate(
            debit_sum=Sum('debit'),
            credit_sum=Sum('credit')
        )
        debit = sums['debit_sum'] or 0
        credit = sums['credit_sum'] or 0

        if str(acc.code).startswith('4'):  # إيرادات
            net_income += (credit - debit)
        elif str(acc.code).startswith('5'):  # مصاريف
            net_income -= (debit - credit)

    # ✅ تجميع الأصول/الخصوم/حقوق الملكية
    for acc in accounts:
        sums = JournalLine.objects.filter(account=acc).aggregate(
            debit_sum=Sum('debit'),
            credit_sum=Sum('credit')
        )
        debit = sums['debit_sum'] or 0
        credit = sums['credit_sum'] or 0

        if str(acc.code).startswith('1'):  # أصول
            balance = debit - credit
            if balance != 0:
                assets.append({'code': acc.code, 'name': acc.name, 'balance': balance})
                total_assets += balance

        elif str(acc.code).startswith('2'):  # خصوم
            balance = credit - debit
            if balance != 0:
                liabilities.append({'code': acc.code, 'name': acc.name, 'balance': balance})
                total_liabilities += balance

        elif str(acc.code).startswith('3'):  # حقوق ملكية
            balance = credit - debit
            if balance != 0:
                equity.append({'code': acc.code, 'name': acc.name, 'balance': balance})
                total_equity += balance

    # ✅ أضف صافي الربح لحقوق الملكية عشان تتوازن الميزانية
    if net_income != 0:
        equity.append({'code': '-', 'name': 'صافي الربح/الخسارة', 'balance': net_income})
        total_equity += net_income

    is_balanced = (round(float(total_assets), 2) == round(float(total_liabilities + total_equity), 2))

    context = {
        'assets': assets,
        'liabilities': liabilities,
        'equity': equity,
        'total_assets': total_assets,
        'total_liabilities': total_liabilities,
        'total_equity': total_equity,
        'is_balanced': is_balanced,
    }
    return render(request, 'accounting_app/balance_sheet.html', context)


from django.db.models import Sum
from .models import JournalLine, Account


@login_required
def trial_balance(request):
    period_id = (request.GET.get("period") or "").strip()
    periods = AccountingPeriod.objects.all().order_by("-start_date")
    accounts = Account.objects.all().order_by("code")

    selected_period = None
    if period_id:
        selected_period = AccountingPeriod.objects.filter(id=period_id).first()

    rows = []

    # Totals
    total_opening_debit = 0.0
    total_opening_credit = 0.0
    total_move_debit = 0.0
    total_move_credit = 0.0
    total_closing_debit = 0.0
    total_closing_credit = 0.0

    for acc in accounts:
        base = JournalLine.objects.filter(account=acc)

        # Opening = قبل بداية الفترة (إذا في فترة)
        opening_bal = 0.0
        if selected_period:
            o = base.filter(entry__date__lt=selected_period.start_date).aggregate(
                d=Sum("debit"), c=Sum("credit")
            )
            opening_bal = float(o["d"] or 0) - float(o["c"] or 0)

        # Movement = داخل الفترة (أو كل شيء إذا ما في فترة)
        move_qs = base
        if selected_period:
            move_qs = move_qs.filter(
                entry__date__gte=selected_period.start_date,
                entry__date__lte=selected_period.end_date,
            )

        m = move_qs.aggregate(d=Sum("debit"), c=Sum("credit"))
        move_debit = float(m["d"] or 0)
        move_credit = float(m["c"] or 0)

        closing_bal = opening_bal + (move_debit - move_credit)

        # تجاهل الحسابات اللي صفر بالكامل
        if opening_bal == 0 and move_debit == 0 and move_credit == 0 and closing_bal == 0:
            continue

        # تحويل opening/closing إلى (مدين/دائن)
        opening_debit = opening_bal if opening_bal > 0 else 0.0
        opening_credit = abs(opening_bal) if opening_bal < 0 else 0.0

        closing_debit = closing_bal if closing_bal > 0 else 0.0
        closing_credit = abs(closing_bal) if closing_bal < 0 else 0.0

        total_opening_debit += opening_debit
        total_opening_credit += opening_credit
        total_move_debit += move_debit
        total_move_credit += move_credit
        total_closing_debit += closing_debit
        total_closing_credit += closing_credit

        rows.append({
            "code": acc.code,
            "name": acc.name,
            "opening_debit": opening_debit,
            "opening_credit": opening_credit,
            "move_debit": move_debit,
            "move_credit": move_credit,
            "closing_debit": closing_debit,
            "closing_credit": closing_credit,
        })

    context = {
        "periods": periods,
        "selected_period": period_id,
        "rows": rows,

        "total_opening_debit": f"{total_opening_debit:.2f}",
        "total_opening_credit": f"{total_opening_credit:.2f}",
        "total_move_debit": f"{total_move_debit:.2f}",
        "total_move_credit": f"{total_move_credit:.2f}",
        "total_closing_debit": f"{total_closing_debit:.2f}",
        "total_closing_credit": f"{total_closing_credit:.2f}",
    }
    return render(request, "accounting_app/trial_balance.html", context)

from django.db.models import Sum

@login_required
def cash_management(request):
    period_id = (request.GET.get("period") or "").strip()
    account_id = (request.GET.get("account") or "").strip()

    periods = AccountingPeriod.objects.all().order_by("-start_date")
    accounts = Account.objects.all().order_by("code")

    selected_period = None
    if period_id:
        selected_period = AccountingPeriod.objects.filter(id=period_id).first()

    # بدون اختيار حساب: نعرض بس الفلاتر
    if not account_id:
        return render(request, "accounting_app/cash_management.html", {
            "periods": periods,
            "accounts": accounts,
            "selected_period": period_id,
            "selected_account": "",
            "rows": [],
            "opening_text": "0.00",
            "opening_class": "",
            "receipts": "0.00",
            "payments": "0.00",
            "closing_text": "0.00",
            "closing_class": "",
            "hint": "اختاري حساب الصندوق/البنك لعرض حركة النقدية.",
        })

    # Helpers
    def fmt_balance(bal: float):
        if bal > 0:
            return f"{bal:.2f} مدين", "text-success fw-bold"
        if bal < 0:
            return f"{abs(bal):.2f} دائن", "text-danger fw-bold"
        return "0.00", ""

    base = JournalLine.objects.select_related("entry", "account").filter(account_id=account_id)

    # Opening
    opening = 0.0
    if selected_period:
        o = base.filter(entry__date__lt=selected_period.start_date).aggregate(
            d=Sum("debit"), c=Sum("credit")
        )
        opening = float(o["d"] or 0) - float(o["c"] or 0)

        lines = base.filter(
            entry__date__gte=selected_period.start_date,
            entry__date__lte=selected_period.end_date
        ).order_by("entry__date", "entry__id", "id")
    else:
        lines = base.order_by("entry__date", "entry__id", "id")

    # Receipts / Payments within selected range
    sums = lines.aggregate(d=Sum("debit"), c=Sum("credit"))
    receipts = float(sums["d"] or 0)
    payments = float(sums["c"] or 0)

    running = opening
    rows = []

    for line in lines:
        debit = float(line.debit or 0)
        credit = float(line.credit or 0)
        running += (debit - credit)

        bal_text, bal_class = fmt_balance(running)

        entry = line.entry
        rows.append({
            "date": entry.date,
            "entry_serial": entry.serial_number or entry.id,
            "entry_id": entry.id,
            "reference": entry.reference or "-",
            "description": entry.description,
            "note": line.note or "",
            "receipt": debit,   # مدين على النقدية = قبض
            "payment": credit,  # دائن على النقدية = دفع
            "balance_text": bal_text,
            "balance_class": bal_class,
        })

    opening_text, opening_class = fmt_balance(opening)
    closing_text, closing_class = fmt_balance(running)

    return render(request, "accounting_app/cash_management.html", {
        "periods": periods,
        "accounts": accounts,
        "selected_period": period_id,
        "selected_account": account_id,
        "rows": rows,
        "opening_text": opening_text,
        "opening_class": opening_class,
        "receipts": f"{receipts:.2f}",
        "payments": f"{payments:.2f}",
        "closing_text": closing_text,
        "closing_class": closing_class,
        "hint": "",
    })


@login_required
def add_account(request):
    if request.method == "POST":
        form = AccountForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "تم إضافة الحساب بنجاح")
            return redirect("account:chart_of_accounts")
    else:
        form = AccountForm()
    return render(request, "accounting_app/add_account.html", {"form": form})


# ===============================
# إضافة حساب فرعي
# ===============================
@login_required
def add_subaccount(request, parent_id):
    parent = Account.objects.get(id=parent_id)
    if request.method == "POST":
        form = AccountForm(request.POST)
        if form.is_valid():
            subaccount = form.save(commit=False)
            subaccount.parent = parent
            subaccount.save()
            messages.success(request, "تم إضافة الحساب الفرعي بنجاح")
            return redirect("account:chart_of_accounts")
    else:
        form = AccountForm()
    return render(request, "accounting_app/add_account.html", {"form": form, "parent": parent})


# ===============================
# تعديل الحساب
# ===============================
@login_required
def edit_account(request, account_id):
    account = Account.objects.get(id=account_id)
    if request.method == "POST":
        form = AccountForm(request.POST, instance=account)
        if form.is_valid():
            form.save()
            messages.success(request, "تم تعديل الحساب بنجاح")
            return redirect("account:chart_of_accounts")
    else:
        form = AccountForm(instance=account)
    return render(request, "accounting_app/add_account.html", {"form": form, "edit": True})


# ===============================
# حذف الحساب
# ===============================
@login_required
def delete_account(request, account_id):
    account = Account.objects.get(id=account_id)
    account.delete()
    messages.success(request, "تم حذف الحساب بنجاح")
    return redirect("account:chart_of_accounts")

@login_required
def opening_balances(request):
    period_id = (request.GET.get("period") or "").strip()
    periods = AccountingPeriod.objects.all().order_by("-start_date")

    if not period_id:
        return render(request, "accounting_app/opening_balances.html", {
            "periods": periods,
            "selected_period": "",
            "formset": None,
            "hint": "اختاري فترة لإدخال الأرصدة الافتتاحية."
        })

    period = get_object_or_404(AccountingPeriod, id=period_id)

    # منع التعديل إذا الفترة مقفلة
    readonly = bool(period.is_closed)

    # جهزي سطور افتتاحي للحسابات (نخليها للأصول/الخصوم/الحقوق فقط)
    base_accounts = Account.objects.exclude(account_type__in=[Account.REVENUE, Account.EXPENSE]).order_by("code")
    for acc in base_accounts:
        OpeningBalance.objects.get_or_create(period=period, account=acc, defaults={"debit": 0, "credit": 0})

    qs = OpeningBalance.objects.select_related("account").filter(period=period).order_by("account__code")

    if request.method == "POST":
        if readonly:
            messages.error(request, "لا يمكن الحفظ: الفترة مقفلة.")
            return redirect(f"{request.path}?period={period.id}")

        formset = OpeningBalanceFormSet(request.POST, queryset=qs)
        if formset.is_valid():
            objs = formset.save(commit=False)
            for obj in objs:
                obj.period = period
                obj.save()
            for obj in formset.deleted_objects:
                obj.delete()

            messages.success(request, f"تم حفظ الافتتاحي للفترة {period.name}")
            return redirect(f"{request.path}?period={period.id}")
    else:
        formset = OpeningBalanceFormSet(queryset=qs)

    return render(request, "accounting_app/opening_balances.html", {
        "periods": periods,
        "selected_period": str(period.id),
        "period": period,
        "formset": formset,
        "readonly": readonly,
        "hint": "",
    })

@login_required
@require_POST
@transaction.atomic
def post_opening_to_journal(request, period_id):
    period = get_object_or_404(AccountingPeriod, id=period_id)

    if period.is_closed:
        messages.error(request, "لا يمكن إنشاء قيد افتتاحي: الفترة مقفلة.")
        return redirect(f"/account/opening-balances/?period={period.id}")

    # منع التكرار: إذا موجود قيد افتتاحي مسبقًا (نميّزه بالمرجع)
    ref = f"OPEN-{period.name}"
    if JournalEntry.objects.filter(period=period, reference=ref).exists():
        messages.warning(request, "تم إنشاء قيد افتتاحي لهذه الفترة مسبقًا.")
        return redirect(f"/account/opening-balances/?period={period.id}")

    obs = OpeningBalance.objects.select_related("account").filter(period=period)
    obs = [ob for ob in obs if (ob.debit or 0) != 0 or (ob.credit or 0) != 0]

    if not obs:
        messages.error(request, "لا يوجد أرصدة افتتاحية بقيم (مدين/دائن) لإنشاء القيد.")
        return redirect(f"/account/opening-balances/?period={period.id}")

    je = JournalEntry.objects.create(
        period=period,
        date=period.start_date,
        reference=ref,
        description=f"قيد افتتاحي للفترة {period.name}",
        created_by=request.user,
    )

    total_d = 0
    total_c = 0

    for ob in obs:
        d = ob.debit or 0
        c = ob.credit or 0
        JournalLine.objects.create(
            entry=je,
            account=ob.account,
            debit=d,
            credit=c,
            note=ob.note or "افتتاحي"
        )
        total_d += float(d)
        total_c += float(c)

    if round(total_d, 2) != round(total_c, 2):
        raise ValidationError("قيد الافتتاحي غير متوازن. تأكدي أن مجموع المدين = مجموع الدائن.")

    messages.success(request, f"تم إنشاء قيد افتتاحي ({je.serial_number or je.id}) للفترة {period.name}")
    return redirect("account:journal_entries")



@login_required
@require_POST
@transaction.atomic
def close_period(request, period_id):
    period = get_object_or_404(AccountingPeriod, id=period_id)

    if period.is_closed:
        messages.info(request, "الفترة مقفلة مسبقًا.")
        return redirect("account:opening_balances")

    # لازم config موجود + retained موجود
    cfg = AccountingConfig.get_config()
    if not cfg.retained_earnings_account_id:
        messages.error(request, "لا يوجد حساب الأرباح المرحلة داخل AccountingConfig.")
        return redirect("account:opening_balances")

    # اجمع الإيرادات/المصاريف داخل الفترة من القيود
    rev_accounts = Account.objects.filter(account_type=Account.REVENUE)
    exp_accounts = Account.objects.filter(account_type=Account.EXPENSE)

    def sums_for(acc):
        s = JournalLine.objects.filter(
            account=acc,
            entry__date__gte=period.start_date,
            entry__date__lte=period.end_date,
        ).aggregate(d=Sum("debit"), c=Sum("credit"))
        d = float(s["d"] or 0)
        c = float(s["c"] or 0)
        return d, c

    # قيد الإقفال بتاريخ نهاية الفترة
    je = JournalEntry.objects.create(
        period=period,
        date=period.end_date,
        reference=f"CLOSE-{period.name}",
        description=f"قيد إقفال الفترة {period.name}",
        created_by=request.user,
    )

    total_income = 0.0

    # الإيرادات: عادةً رصيدها دائن => لإقفالها نعمل (مدين الإيراد) بقيمة صافيها
    for acc in rev_accounts:
        d, c = sums_for(acc)
        net = c - d  # صافي الإيراد
        if net <= 0:
            continue
        JournalLine.objects.create(entry=je, account=acc, debit=net, credit=0, note="إقفال إيراد")
        total_income += net

    # المصاريف: عادةً رصيدها مدين => لإقفالها نعمل (دائن المصروف) بقيمة صافيها
    total_exp = 0.0
    for acc in exp_accounts:
        d, c = sums_for(acc)
        net = d - c  # صافي المصروف
        if net <= 0:
            continue
        JournalLine.objects.create(entry=je, account=acc, debit=0, credit=net, note="إقفال مصروف")
        total_exp += net

    net_profit = total_income - total_exp

    # إقفال على retained earnings:
    # لو ربح => retained دائن
    # لو خسارة => retained مدين
    if net_profit > 0:
        JournalLine.objects.create(
            entry=je, account=cfg.retained_earnings_account,
            debit=0, credit=net_profit, note="ترحيل صافي الربح للأرباح المرحلة"
        )
    elif net_profit < 0:
        JournalLine.objects.create(
            entry=je, account=cfg.retained_earnings_account,
            debit=abs(net_profit), credit=0, note="ترحيل صافي الخسارة للأرباح المرحلة"
        )
    else:
        # صفر: نلغي القيد إذا ما فيه سطور
        if not je.lines.exists():
            je.delete()

    period.is_closed = True
    period.save(update_fields=["is_closed"])

    messages.success(request, f"تم إنشاء قيد الإقفال وإقفال الفترة {period.name}.")
    return redirect(f"/account/opening-balances/?period={period.id}")

from django.contrib.admin.views.decorators import staff_member_required

@staff_member_required
@require_POST
@transaction.atomic
def reopen_period(request, period_id):
    period = get_object_or_404(AccountingPeriod, id=period_id)

    # فتح الفترة
    period.is_closed = False
    period.save(update_fields=["is_closed"])

    messages.success(request, f"تمت إعادة فتح الفترة {period.name}")
    return redirect(f"/account/opening-balances/?period={period.id}")


@login_required
def payments(request):
    if request.method == "POST":
        form = PaymentForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    p = form.save()
                    p.post_to_journal(user=request.user)
                messages.success(request, "تم حفظ السند وترحيله إلى القيود.")
                return redirect("account:payments")
            except ValidationError as e:
                messages.error(request, str(e))
        else:
            messages.error(request, "يرجى التأكد من بيانات السند.")
    else:
        form = PaymentForm()

    payments_qs = (
        Payment.objects
        .select_related("customer", "supplier", "journal_entry", "cash_account")
        .order_by("-date", "-id")
    )
    return render(request, "accounting_app/payments.html", {"form": form, "payments": payments_qs})



@login_required
def payment_print(request, pk):
    payment = get_object_or_404(Payment, pk=pk)

    title = "سند قبض" if payment.payment_type == Payment.RECEIPT else "سند صرف"
    party_label = "العميل" if payment.payment_type == Payment.RECEIPT else "المورد"
    party_name = payment.customer.name if payment.customer_id else (payment.supplier.name if payment.supplier_id else "-")

    return render(request, "accounting_app/payment_print.html", {
        "payment": payment,
        "title": title,
        "party_label": party_label,
        "party_name": party_name,
    })


@login_required
def payment_pdf(request, pk):
    payment = get_object_or_404(Payment.objects.select_related("customer", "supplier", "journal_entry"), pk=pk)

    font_name = _register_arabic_font()

    company_name = "شركة المحبة للصناعات الغذائية"
    company_address = "الأردن - عمّان - أبو علندا"
    now_str = timezone.localtime(timezone.now()).strftime("%Y-%m-%d %H:%M")

    title = "سند قبض" if payment.payment_type == Payment.RECEIPT else "سند صرف"
    party_label = "العميل" if payment.payment_type == Payment.RECEIPT else "المورد"
    party_name = payment.customer.name if payment.customer_id else (payment.supplier.name if payment.supplier_id else "-")

    # PDF setup
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=24, rightMargin=24, topMargin=70, bottomMargin=45
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title_ar", parent=styles["Title"], fontName=font_name, fontSize=18, alignment=1)
    small_style = ParagraphStyle("small_ar", parent=styles["Normal"], fontName=font_name, fontSize=11, alignment=1)
    cell_style = ParagraphStyle("cell_ar", parent=styles["Normal"], fontName=font_name, fontSize=11, alignment=1)

    def header_footer(canvas, doc_):
        canvas.saveState()
        canvas.setFont(font_name, 12)
        canvas.drawCentredString(doc_.pagesize[0] / 2, doc_.pagesize[1] - 25, _ar(company_name))
        canvas.setFont(font_name, 10)
        canvas.drawCentredString(doc_.pagesize[0] / 2, doc_.pagesize[1] - 42, _ar(company_address))
        canvas.drawCentredString(doc_.pagesize[0] / 2, doc_.pagesize[1] - 58, _ar(f"تاريخ الطباعة: {now_str}"))
        canvas.setFont(font_name, 10)
        canvas.drawRightString(doc_.pagesize[0] - 24, 18, _ar(f"صفحة {doc_.page}"))
        canvas.restoreState()

    elements = []

    elements.append(Paragraph(_ar(title), title_style))
    elements.append(Spacer(1, 8))

    # معلومات السند (سطرين مرتب)
    je_txt = "-"
    if payment.journal_entry_id:
        je_txt = payment.journal_entry.serial_number or str(payment.journal_entry_id)

    elements.append(Paragraph(_ar(
       f"رقم السند: {payment.voucher_number or payment.id} | التاريخ: {payment.date} | {party_label}: {party_name}"

    ), small_style))
    elements.append(Paragraph(_ar(
        f"المبلغ: {payment.amount} | القيد المرتبط: {je_txt}"
    ), small_style))
    cash_name = payment.cash_account.name if payment.cash_account_id else "-"
    elements.append(Paragraph(_ar(f"حساب الصندوق/البنك: {cash_name}"), small_style))

    if payment.note:
        elements.append(Paragraph(_ar(f"ملاحظة: {payment.note}"), small_style))

    elements.append(Spacer(1, 12))

    # جدول القيد المرتبط (إن وجد)
    if payment.journal_entry_id:
        entry = payment.journal_entry
        lines = list(entry.lines.select_related("account").all()) if hasattr(entry, "lines") else []

        headers = ["#", "رمز الحساب", "اسم الحساب", "ملاحظة", "مدين", "دائن"]
        data = [[Paragraph(_ar(h), cell_style) for h in headers]]

        total_debit = 0.0
        total_credit = 0.0

        for i, line in enumerate(lines, start=1):
            debit = float(line.debit or 0)
            credit = float(line.credit or 0)
            total_debit += debit
            total_credit += credit

            data.append([
                Paragraph(_ar(i), cell_style),
                Paragraph(_ar(getattr(line.account, "code", "")), cell_style),
                Paragraph(_ar(getattr(line.account, "name", "")), cell_style),
                Paragraph(_ar(line.note or ""), cell_style),
                Paragraph(_ar(f"{debit:.2f}" if debit else "-"), cell_style),
                Paragraph(_ar(f"{credit:.2f}" if credit else "-"), cell_style),
            ])

        # إجمالي
        data.append([
            Paragraph(_ar(""), cell_style),
            Paragraph(_ar(""), cell_style),
            Paragraph(_ar(""), cell_style),
            Paragraph(_ar("الإجمالي"), cell_style),
            Paragraph(_ar(f"{total_debit:.2f}"), cell_style),
            Paragraph(_ar(f"{total_credit:.2f}"), cell_style),
        ])

        page_w = landscape(A4)[0]
        table_w = page_w - 24 - 24
        col_ratios = [0.06, 0.14, 0.22, 0.30, 0.14, 0.14]
        col_widths = [table_w * r for r in col_ratios]

        table = Table(data, colWidths=col_widths, repeatRows=1, hAlign="CENTER")
        table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, -1), 11),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E6E6E6")),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.whitesmoke, colors.white]),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F1F1F1")),
        ]))

        elements.append(table)
    else:
        elements.append(Paragraph(_ar("لا يوجد قيد مرتبط بهذا السند (غير مرحّل)."), small_style))

    doc.build(elements, onFirstPage=header_footer, onLaterPages=header_footer)
    buffer.seek(0)

    filename = f"payment_{payment.id}.pdf"
    return FileResponse(buffer, as_attachment=True, filename=filename)

from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.utils import timezone
from django.http import FileResponse
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import io
from datetime import datetime

# --- helpers ---
def _parse_date(s):
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None

def _payments_report_qs(payment_type, date_from=None, date_to=None, party_id=None, cash_account_id=None):
    qs = Payment.objects.select_related("customer", "supplier", "cash_account", "journal_entry")\
        .filter(payment_type=payment_type).order_by("-date", "-id")
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)
    if cash_account_id:
        qs = qs.filter(cash_account_id=cash_account_id)
    if party_id:
        if payment_type == Payment.RECEIPT:
            qs = qs.filter(customer_id=party_id)
        else:
            qs = qs.filter(supplier_id=party_id)
    return qs

def _je_ref(je):
    # آمن لكل اختلافات الموديل
    if not je:
        return ""
    return getattr(je, "serial_number", None) or getattr(je, "entry_number", None) or str(getattr(je, "id", ""))

# -------------------------
# Views (Pages)
# -------------------------
@login_required
def receipts_report(request):
    date_from = _parse_date(request.GET.get("from"))
    date_to = _parse_date(request.GET.get("to"))
    party_id = request.GET.get("customer")
    cash_id = request.GET.get("cash_account")

    qs = _payments_report_qs(Payment.RECEIPT, date_from, date_to, party_id, cash_id)
    total = qs.aggregate(s=Sum("amount")).get("s") or 0

    return render(request, "accounting_app/payments_receipts_report.html", {
        "rows": qs,
        "total": total,
        "date_from": date_from,
        "date_to": date_to,
        "customers": Customer.objects.all().order_by("name"),
        "cash_accounts": Account.objects.all().order_by("code"),
        "selected_customer": int(party_id) if party_id else None,
        "selected_cash": int(cash_id) if cash_id else None,
    })

@login_required
def disbursements_report(request):
    date_from = _parse_date(request.GET.get("from"))
    date_to = _parse_date(request.GET.get("to"))
    party_id = request.GET.get("supplier")
    cash_id = request.GET.get("cash_account")

    qs = _payments_report_qs(Payment.DISBURSE, date_from, date_to, party_id, cash_id)
    total = qs.aggregate(s=Sum("amount")).get("s") or 0

    return render(request, "accounting_app/payments_disbursements_report.html", {
        "rows": qs,
        "total": total,
        "date_from": date_from,
        "date_to": date_to,
        "suppliers": Supplier.objects.all().order_by("name"),
        "cash_accounts": Account.objects.all().order_by("code"),
        "selected_supplier": int(party_id) if party_id else None,
        "selected_cash": int(cash_id) if cash_id else None,
    })

# -------------------------
# Excel (uses your improved function)
# -------------------------
@login_required
def receipts_report_excel(request):
    date_from = _parse_date(request.GET.get("from"))
    date_to = _parse_date(request.GET.get("to"))
    party_id = request.GET.get("customer")
    cash_id = request.GET.get("cash_account")

    qs = _payments_report_qs(Payment.RECEIPT, date_from, date_to, party_id, cash_id)

    cust_name = ""
    if party_id:
        c = Customer.objects.filter(id=party_id).first()
        cust_name = c.name if c else ""

    cash_name = ""
    if cash_id:
        a = Account.objects.filter(id=cash_id).first()
        cash_name = str(a) if a else ""

    filters_line = f"من: {date_from or '-'} | إلى: {date_to or '-'} | العميل: {cust_name or '-'} | الحساب: {cash_name or '-'}"
    return _export_payments_excel("receipts_report.xlsx", "تقرير سندات القبض", qs, "قبض", filters_line=filters_line)

@login_required
def disbursements_report_excel(request):
    date_from = _parse_date(request.GET.get("from"))
    date_to = _parse_date(request.GET.get("to"))
    party_id = request.GET.get("supplier")
    cash_id = request.GET.get("cash_account")

    qs = _payments_report_qs(Payment.DISBURSE, date_from, date_to, party_id, cash_id)

    sup_name = ""
    if party_id:
        s = Supplier.objects.filter(id=party_id).first()
        sup_name = s.name if s else ""

    cash_name = ""
    if cash_id:
        a = Account.objects.filter(id=cash_id).first()
        cash_name = str(a) if a else ""

    filters_line = f"من: {date_from or '-'} | إلى: {date_to or '-'} | المورد: {sup_name or '-'} | الحساب: {cash_name or '-'}"
    return _export_payments_excel("disbursements_report.xlsx", "تقرير سندات الصرف", qs, "صرف", filters_line=filters_line)

# -------------------------
# PDF (uses your improved function)
# -------------------------
@login_required
def receipts_report_pdf(request):
    date_from = _parse_date(request.GET.get("from"))
    date_to = _parse_date(request.GET.get("to"))
    party_id = request.GET.get("customer")
    cash_id = request.GET.get("cash_account")

    qs = _payments_report_qs(Payment.RECEIPT, date_from, date_to, party_id, cash_id)

    cust_name = ""
    if party_id:
        c = Customer.objects.filter(id=party_id).first()
        cust_name = c.name if c else ""

    cash_name = ""
    if cash_id:
        a = Account.objects.filter(id=cash_id).first()
        cash_name = str(a) if a else ""

    filters_line = f"من: {date_from or '-'} | إلى: {date_to or '-'} | العميل: {cust_name or '-'} | الحساب: {cash_name or '-'}"
    return _export_payments_pdf("receipts_report.pdf", "تقرير سندات القبض", qs, "قبض", filters_line=filters_line)

@login_required
def disbursements_report_pdf(request):
    date_from = _parse_date(request.GET.get("from"))
    date_to = _parse_date(request.GET.get("to"))
    party_id = request.GET.get("supplier")
    cash_id = request.GET.get("cash_account")

    qs = _payments_report_qs(Payment.DISBURSE, date_from, date_to, party_id, cash_id)

    sup_name = ""
    if party_id:
        s = Supplier.objects.filter(id=party_id).first()
        sup_name = s.name if s else ""

    cash_name = ""
    if cash_id:
        a = Account.objects.filter(id=cash_id).first()
        cash_name = str(a) if a else ""

    filters_line = f"من: {date_from or '-'} | إلى: {date_to or '-'} | المورد: {sup_name or '-'} | الحساب: {cash_name or '-'}"
    return _export_payments_pdf("disbursements_report.pdf", "تقرير سندات الصرف", qs, "صرف", filters_line=filters_line)


from decimal import Decimal
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# -------- Helpers for Excel --------
def _export_payments_excel(filename, title, qs):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.sheet_view.rightToLeft = True

    # Header title
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:G1")

    now_str = timezone.localtime(timezone.now()).strftime("%Y-%m-%d %H:%M")
    ws["A2"] = f"تاريخ الطباعة: {now_str}"
    ws.merge_cells("A2:G2")

    headers = ["#", "رقم السند", "التاريخ", "الطرف", "حساب الصندوق/البنك", "المبلغ", "القيد"]
    ws.append([])
    ws.append(headers)

    for c in range(1, 8):
        ws.cell(row=4, column=c).font = Font(bold=True)
        ws.cell(row=4, column=c).alignment = Alignment(horizontal="center", vertical="center")

    total = Decimal("0")
    row_i = 5
    for i, p in enumerate(qs, start=1):
        party = p.customer.name if p.customer_id else (p.supplier.name if p.supplier_id else "")
        cash = p.cash_account.name if p.cash_account_id else ""
        je = p.journal_entry.serial_number if p.journal_entry_id else "—"
        ws.append([
            i,
            p.voucher_number or p.id,
            p.date.strftime("%Y-%m-%d"),
            party,
            cash,
            float(p.amount),
            je
        ])
        total += (p.amount or 0)
        row_i += 1

    ws.append(["", "", "", "", "الإجمالي", float(total), ""])

    # alignment + widths
    for r in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=7):
        for cell in r:
            if cell.column in (6,):  # amount
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = [6, 22, 14, 28, 28, 14, 22]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp


# -------- PDF report (your function) --------
def _export_payments_pdf(filename, title, qs):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=24, leftMargin=24, topMargin=36, bottomMargin=24)
    font_name = _register_arabic_font()

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "title",
        parent=styles["Title"],
        fontName=font_name,
        fontSize=16,
        alignment=1,
        spaceAfter=12,
    )
    small = ParagraphStyle(
        "small",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=10,
        alignment=2,
    )
    cell = ParagraphStyle(
        "cell",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=10,
        alignment=1,
        leading=12,
    )

    elements = []
    elements.append(Paragraph(_ar(title), title_style))
    elements.append(Paragraph(_ar(f"تاريخ الطباعة: {timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M')}"), small))
    elements.append(Spacer(1, 10))

    data = [[Paragraph(_ar(h), cell) for h in ["#", "رقم السند", "التاريخ", "الطرف", "حساب الصندوق/البنك", "المبلغ", "القيد"]]]

    total = Decimal("0")
    for i, p in enumerate(qs, start=1):
        party = p.customer.name if p.customer_id else (p.supplier.name if p.supplier_id else "")
        cash = p.cash_account.name if p.cash_account_id else ""
        je = p.journal_entry.serial_number if p.journal_entry_id else "—"
        data.append([
            Paragraph(_ar(i), cell),
            Paragraph(_ar(p.voucher_number or "—"), cell),
            Paragraph(_ar(p.date.strftime("%Y-%m-%d")), cell),
            Paragraph(_ar(party), cell),
            Paragraph(_ar(cash), cell),
            Paragraph(_ar(f"{p.amount:,.2f}"), cell),
            Paragraph(_ar(je), cell),
        ])
        total += (p.amount or 0)

    data.append([
        Paragraph(_ar(""), cell),
        Paragraph(_ar(""), cell),
        Paragraph(_ar(""), cell),
        Paragraph(_ar(""), cell),
        Paragraph(_ar("الإجمالي"), cell),
        Paragraph(_ar(f"{total:,.2f}"), cell),
        Paragraph(_ar(""), cell),
    ])

    table = Table(data, repeatRows=1, colWidths=[28, 90, 65, 120, 130, 70, 60], hAlign="CENTER")
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EFEFEF")),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.whitesmoke, colors.white]),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F7F7F7")),
        ("SPAN", (0, -1), (4, -1)),
        ("ALIGN", (5, 1), (5, -1), "RIGHT"),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename=filename)


# -------- Report views --------
@login_required
def receipts_report(request):
    qs = (Payment.objects
          .select_related("customer", "supplier", "journal_entry", "cash_account")
          .filter(payment_type=Payment.RECEIPT)
          .order_by("-date", "-id"))

    export = (request.GET.get("export") or "").lower()
    if export == "pdf":
        return _export_payments_pdf("receipts_report.pdf", "تقرير سندات القبض", qs)
    if export == "excel":
        return _export_payments_excel("receipts_report.xlsx", "تقرير سندات القبض", qs)

    return render(request, "accounting_app/payments_report.html", {
        "title": "تقرير سندات القبض",
        "rows": qs,
        "kind": "receipts",
    })


@login_required
def disbursements_report(request):
    qs = (Payment.objects
          .select_related("customer", "supplier", "journal_entry", "cash_account")
          .filter(payment_type=Payment.DISBURSE)
          .order_by("-date", "-id"))

    export = (request.GET.get("export") or "").lower()
    if export == "pdf":
        return _export_payments_pdf("disbursements_report.pdf", "تقرير سندات الصرف", qs)
    if export == "excel":
        return _export_payments_excel("disbursements_report.xlsx", "تقرير سندات الصرف", qs)

    return render(request, "accounting_app/payments_report.html", {
        "title": "تقرير سندات الصرف",
        "rows": qs,
        "kind": "disbursements",
    })

from datetime import date
from django.http import HttpResponse
from django.shortcuts import render
from django.contrib.auth.decorators import login_required

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm

from accounting_app.models import SalesInvoice, PurchaseInvoice


def _parse_date(s: str | None) -> date | None:
    if not s:
        return None
    try:
        return date.fromisoformat(s)  # expects YYYY-MM-DD
    except ValueError:
        return None


def _invoices_pdf_response(title: str, rows: list[dict], filename: str) -> HttpResponse:
    """
    rows: list of dicts with keys:
      number, inv_date, party, total, posted
    """
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    c = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # Header
    y = height - 2 * cm
    c.setFont("Helvetica-Bold", 14)
    c.drawRightString(width - 2 * cm, y, title)

    y -= 0.8 * cm
    c.setFont("Helvetica", 10)
    c.drawRightString(width - 2 * cm, y, f"عدد الفواتير: {len(rows)}")

    # Table header (RTL-friendly layout)
    y -= 1.2 * cm
    c.setFont("Helvetica-Bold", 10)

    # Column positions (from right to left)
    x_right = width - 2 * cm
    col_w = {
        "posted": 2.0 * cm,
        "total": 3.0 * cm,
        "party": 6.5 * cm,
        "date": 3.0 * cm,
        "number": 3.0 * cm,
    }

    def draw_row(ypos: float, row: dict, bold: bool = False):
        c.setFont("Helvetica-Bold" if bold else "Helvetica", 9)
        x = x_right

        # Posted
        c.drawRightString(x, ypos, row["posted"])
        x -= col_w["posted"]

        # Total
        c.drawRightString(x, ypos, row["total"])
        x -= col_w["total"]

        # Party
        c.drawRightString(x, ypos, row["party"])
        x -= col_w["party"]

        # Date
        c.drawRightString(x, ypos, row["inv_date"])
        x -= col_w["date"]

        # Number
        c.drawRightString(x, ypos, row["number"])

    # Header labels
    header = {
        "posted": "مرحّلة",
        "total": "الإجمالي",
        "party": "العميل/المورد",
        "inv_date": "التاريخ",
        "number": "رقم الفاتورة",
    }
    draw_row(y, {
        "posted": header["posted"],
        "total": header["total"],
        "party": header["party"],
        "inv_date": header["inv_date"],
        "number": header["number"],
    }, bold=True)

    y -= 0.45 * cm
    c.setLineWidth(0.5)
    c.line(2 * cm, y, width - 2 * cm, y)

    # Table rows
    y -= 0.6 * cm
    c.setFont("Helvetica", 9)

    for r in rows:
        if y < 2.5 * cm:
            c.showPage()
            y = height - 2 * cm
            c.setFont("Helvetica-Bold", 14)
            c.drawRightString(width - 2 * cm, y, title)
            y -= 1.2 * cm
            c.setFont("Helvetica-Bold", 10)
            # Repeat header
            draw_row(y, {
                "posted": header["posted"],
                "total": header["total"],
                "party": header["party"],
                "inv_date": header["inv_date"],
                "number": header["number"],
            }, bold=True)
            y -= 0.45 * cm
            c.line(2 * cm, y, width - 2 * cm, y)
            y -= 0.6 * cm

        draw_row(y, r, bold=False)
        y -= 0.5 * cm

    c.showPage()
    c.save()
    return response


@login_required
def sales_invoices_report(request):
    date_from = _parse_date(request.GET.get("date_from"))
    date_to = _parse_date(request.GET.get("date_to"))
    fmt = request.GET.get("format")

    qs = SalesInvoice.objects.all().order_by("-date", "-id")

    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)

    invoices = list(qs)

    # Prepare rows for screen + pdf
    rows = []
    for inv in invoices:
        customer_name = getattr(getattr(inv, "customer", None), "name", "") or "-"
        posted = "نعم" if getattr(inv, "journal_entry_id", None) else "لا"
        rows.append({
            "number": str(getattr(inv, "id", "")),
            "inv_date": str(getattr(inv, "date", "")),
            "party": customer_name,
            "total": str(getattr(inv, "total", "")),
            "posted": posted,
        })

    if fmt == "pdf":
        title = "تقرير فواتير البيع"
        filename = "sales_invoices_report.pdf"
        return _invoices_pdf_response(title, rows, filename)

    return render(request, "accounting_app/sales_invoices_report.html", {
        "invoices": invoices,
        "date_from": request.GET.get("date_from", ""),
        "date_to": request.GET.get("date_to", ""),
    })


@login_required
def purchase_invoices_report(request):
    date_from = _parse_date(request.GET.get("date_from"))
    date_to = _parse_date(request.GET.get("date_to"))
    fmt = request.GET.get("format")

    qs = PurchaseInvoice.objects.all().order_by("-date", "-id")

    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)

    invoices = list(qs)

    rows = []
    for inv in invoices:
        supplier_name = getattr(getattr(inv, "supplier", None), "name", "") or "-"
        posted = "نعم" if getattr(inv, "journal_entry_id", None) else "لا"
        rows.append({
            "number": str(getattr(inv, "id", "")),
            "inv_date": str(getattr(inv, "date", "")),
            "party": supplier_name,
            "total": str(getattr(inv, "total", "")),
            "posted": posted,
        })

    if fmt == "pdf":
        title = "تقرير فواتير الشراء"
        filename = "purchase_invoices_report.pdf"
        return _invoices_pdf_response(title, rows, filename)

    return render(request, "accounting_app/purchase_invoices_report.html", {
        "invoices": invoices,
        "date_from": request.GET.get("date_from", ""),
        "date_to": request.GET.get("date_to", ""),
    })

from django.shortcuts import get_object_or_404
from accounting_app.models import Customer, Supplier, SalesInvoice, PurchaseInvoice, Payment


def _money(x):
    try:
        return f"{x:.2f}"
    except Exception:
        return f"{x}" if x is not None else "0.00"


import io
from datetime import date
from decimal import Decimal

from django.http import HttpResponse, FileResponse
from django.shortcuts import get_object_or_404, render
from django.contrib.auth.decorators import login_required
from django.utils import timezone

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from accounting_app.models import Customer, Supplier, SalesInvoice, PurchaseInvoice, Payment


def _format_balance_with_label(val) -> str:
    """
    يرجع الرصيد بصيغة واضحة:
    40.00 (مدين) أو 40.00 (دائن) أو 0.00
    """
    try:
        x = Decimal(str(val))
    except Exception:
        try:
            x = Decimal(str(float(val)))
        except Exception:
            return str(val)

    if x == 0:
        return "0.00"
    if x > 0:
        return f"{x:,.2f} (مدين)"
    return f"{abs(x):,.2f} (دائن)"


def _period_text(date_from: str, date_to: str) -> str:
    if date_from and date_to:
        return f"من {date_from} إلى {date_to}"
    if date_from:
        return f"من {date_from}"
    if date_to:
        return f"حتى {date_to}"
    return "كل الفترات"


def _statement_pdf(
    title: str,
    party_name: str,
    rows: list[dict],
    filename: str,
    date_from: str = "",
    date_to: str = ""
) -> HttpResponse:
    """
    rows: list of dict keys:
      date, doc, ref, note, debit, credit, balance
    Uses the same Arabic font + Paragraph/Table approach used in other PDFs.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=24,
        leftMargin=24,
        topMargin=30,
        bottomMargin=24
    )
    font_name = _register_arabic_font()

    styles = getSampleStyleSheet()

    company_style = ParagraphStyle(
        "company",
        parent=styles["Title"],
        fontName=font_name,
        fontSize=14,
        alignment=1,
        spaceAfter=2,
    )
    addr_style = ParagraphStyle(
        "addr",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=10,
        alignment=1,
        spaceAfter=8,
    )
    title_style = ParagraphStyle(
        "title",
        parent=styles["Title"],
        fontName=font_name,
        fontSize=16,
        alignment=1,
        spaceAfter=8,
    )
    small = ParagraphStyle(
        "small",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=10,
        alignment=2,  # يمين
        leading=12,
    )
    cell = ParagraphStyle(
        "cell",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=10,
        alignment=1,  # وسط
        leading=12,
    )
    cell_right = ParagraphStyle(
        "cell_right",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=10,
        alignment=2,  # يمين
        leading=12,
    )

    # ✅ عدليهم حسب شركتك
    company_name = "شركة المحبة للصناعات الغذائية"
    company_addr = "الأردن - عمّان - أبو علندا"

    printed_at = timezone.localtime(timezone.now()).strftime("%Y-%m-%d %H:%M")
    period = _period_text(date_from, date_to)

    elements = []
    # Header like previous PDFs
    elements.append(Paragraph(_ar(company_name), company_style))
    elements.append(Paragraph(_ar(company_addr), addr_style))
    elements.append(Paragraph(_ar(title), title_style))

    elements.append(Paragraph(_ar(f"الطرف: {party_name}"), small))
    elements.append(Paragraph(_ar(f"الفترة: {period}"), small))
    elements.append(Paragraph(_ar(f"تاريخ الطباعة: {printed_at}"), small))
    elements.append(Spacer(1, 10))

    headers = ["#", "التاريخ", "النوع", "المرجع", "ملاحظة", "مدين", "دائن", "الرصيد"]
    data = [[Paragraph(_ar(h), cell) for h in headers]]

    total_debit = Decimal("0")
    total_credit = Decimal("0")

    for i, r in enumerate(rows, start=1):
        debit = r.get("debit", "") or ""
        credit = r.get("credit", "") or ""

        # Totals (ignore non-numeric)
        try:
            if debit:
                total_debit += Decimal(str(debit).replace(",", ""))
        except Exception:
            pass
        try:
            if credit:
                total_credit += Decimal(str(credit).replace(",", ""))
        except Exception:
            pass

        data.append([
            Paragraph(_ar(i), cell),
            Paragraph(_ar(r.get("date", "")), cell),
            Paragraph(_ar(r.get("doc", "")), cell),
            Paragraph(_ar(r.get("ref", "")), cell_right),
            Paragraph(_ar(r.get("note", "")), cell_right),
            Paragraph(_ar(debit), cell),
            Paragraph(_ar(credit), cell),
            Paragraph(_ar(r.get("balance", "")), cell),
        ])

    # Totals row
    data.append([
        Paragraph(_ar(""), cell),
        Paragraph(_ar(""), cell),
        Paragraph(_ar(""), cell),
        Paragraph(_ar(""), cell),
        Paragraph(_ar("الإجمالي"), cell),
        Paragraph(_ar(f"{total_debit:,.2f}"), cell),
        Paragraph(_ar(f"{total_credit:,.2f}"), cell),
        Paragraph(_ar(""), cell),
    ])

    table = Table(
        data,
        repeatRows=1,
        # ✅ وسّعنا عمود التاريخ حتى ما ينزل لسطرين
        colWidths=[26, 75, 70, 65, 175, 55, 55, 70],
        hAlign="CENTER"
    )
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 10),

        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EFEFEF")),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),

        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.whitesmoke, colors.white]),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F7F7F7")),
        ("SPAN", (0, -1), (4, -1)),

        # Align numbers
        ("ALIGN", (5, 1), (7, -1), "RIGHT"),

        # Right align ref + note
        ("ALIGN", (3, 1), (4, -2), "RIGHT"),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename=filename)


@login_required
def customer_statement(request, customer_id: int):
    customer = get_object_or_404(Customer, pk=customer_id)

    date_from_obj = _parse_date(request.GET.get("date_from"))
    date_to_obj = _parse_date(request.GET.get("date_to"))
    fmt = request.GET.get("format")

    # raw strings for period text in pdf
    date_from_str = request.GET.get("date_from", "") or ""
    date_to_str = request.GET.get("date_to", "") or ""

    inv_qs = SalesInvoice.objects.filter(customer=customer).order_by("date", "id")

    # Receipts for this customer (قبض)
    pay_qs = (
        Payment.objects.filter(customer=customer, payment_type=Payment.RECEIPT).order_by("date", "id")
        if hasattr(Payment, "RECEIPT")
        else Payment.objects.filter(customer=customer).order_by("date", "id")
    )

    # Opening balance (before date_from)
    opening = 0.0
    if date_from_obj:
        for inv in inv_qs.filter(date__lt=date_from_obj):
            opening += float(getattr(inv, "total", 0) or 0)
        for p in pay_qs.filter(date__lt=date_from_obj):
            opening -= float(getattr(p, "amount", 0) or 0)

    inv_range = inv_qs
    pay_range = pay_qs
    if date_from_obj:
        inv_range = inv_range.filter(date__gte=date_from_obj)
        pay_range = pay_range.filter(date__gte=date_from_obj)
    if date_to_obj:
        inv_range = inv_range.filter(date__lte=date_to_obj)
        pay_range = pay_range.filter(date__lte=date_to_obj)

    movements = []

    for inv in inv_range:
        movements.append({
            "date": getattr(inv, "date", None),
            "doc": "فاتورة بيع",
            "ref": f"SI-{inv.id}",
            "note": getattr(inv, "note", "") or getattr(inv, "description", "") or "",
            "debit": float(getattr(inv, "total", 0) or 0),  # يزيد ذمم العميل (مدين)
            "credit": 0.0,
        })

    for p in pay_range:
        movements.append({
            "date": getattr(p, "date", None),
            "doc": "سند قبض",
            "ref": f"RC-{p.id}",
            "note": getattr(p, "note", "") or "",
            "debit": 0.0,
            "credit": float(getattr(p, "amount", 0) or 0),  # يقلل ذمم العميل
        })

    movements.sort(key=lambda x: (x["date"] or date.min, x["ref"]))

    rows = []
    balance = Decimal(str(opening))

    # opening row
    rows.append({
        "date": "",
        "doc": "",
        "ref": "",
        "note": "رصيد افتتاحي (قبل الفترة)",
        "debit": "",
        "credit": "",
        "balance": _format_balance_with_label(balance),
    })

    for m in movements:
        balance = balance + Decimal(str(m["debit"])) - Decimal(str(m["credit"]))
        rows.append({
            "date": str(m["date"] or ""),
            "doc": m["doc"],
            "ref": m["ref"],
            "note": m["note"],
            "debit": _money(m["debit"]) if m["debit"] else "",
            "credit": _money(m["credit"]) if m["credit"] else "",
            "balance": _format_balance_with_label(balance),
        })

    if fmt == "pdf":
        return _statement_pdf(
            title="كشف حساب عميل",
            party_name=getattr(customer, "name", str(customer)),
            rows=rows,
            filename=f"customer_statement_{customer_id}.pdf",
            date_from=date_from_str,
            date_to=date_to_str,
        )

    return render(request, "accounting_app/customer_statement.html", {
        "customer": customer,
        "rows": rows,
        "date_from": date_from_str,
        "date_to": date_to_str,
    })


@login_required
def supplier_statement(request, supplier_id: int):
    supplier = get_object_or_404(Supplier, pk=supplier_id)

    date_from_obj = _parse_date(request.GET.get("date_from"))
    date_to_obj = _parse_date(request.GET.get("date_to"))
    fmt = request.GET.get("format")

    date_from_str = request.GET.get("date_from", "") or ""
    date_to_str = request.GET.get("date_to", "") or ""

    inv_qs = PurchaseInvoice.objects.filter(supplier=supplier).order_by("date", "id")

    # Disbursements for this supplier (صرف)
    pay_qs = (
        Payment.objects.filter(supplier=supplier, payment_type=Payment.DISBURSE).order_by("date", "id")
        if hasattr(Payment, "DISBURSE")
        else Payment.objects.filter(supplier=supplier).order_by("date", "id")
    )

    # "مستحق للمورد" يزيد مع فاتورة شراء (credit) ويقل مع سند صرف (debit)
    opening = 0.0
    if date_from_obj:
        for inv in inv_qs.filter(date__lt=date_from_obj):
            opening += float(getattr(inv, "total", 0) or 0)
        for p in pay_qs.filter(date__lt=date_from_obj):
            opening -= float(getattr(p, "amount", 0) or 0)

    inv_range = inv_qs
    pay_range = pay_qs
    if date_from_obj:
        inv_range = inv_range.filter(date__gte=date_from_obj)
        pay_range = pay_range.filter(date__gte=date_from_obj)
    if date_to_obj:
        inv_range = inv_range.filter(date__lte=date_to_obj)
        pay_range = pay_range.filter(date__lte=date_to_obj)

    movements = []

    for inv in inv_range:
        movements.append({
            "date": getattr(inv, "date", None),
            "doc": "فاتورة شراء",
            "ref": f"PI-{inv.id}",
            "note": getattr(inv, "note", "") or getattr(inv, "description", "") or "",
            "debit": 0.0,
            "credit": float(getattr(inv, "total", 0) or 0),  # يزيد مستحق المورد
        })

    for p in pay_range:
        movements.append({
            "date": getattr(p, "date", None),
            "doc": "سند صرف",
            "ref": f"PV-{p.id}",
            "note": getattr(p, "note", "") or "",
            "debit": float(getattr(p, "amount", 0) or 0),   # يقلل مستحق المورد
            "credit": 0.0,
        })

    movements.sort(key=lambda x: (x["date"] or date.min, x["ref"]))

    rows = []
    balance = Decimal(str(opening))

    rows.append({
        "date": "",
        "doc": "",
        "ref": "",
        "note": "رصيد افتتاحي (قبل الفترة)",
        "debit": "",
        "credit": "",
        "balance": _format_balance_with_label(balance),
    })

    for m in movements:
        # ✅ نفس منطقك: يزيد مع credit ويقل مع debit
        balance = balance + Decimal(str(m["credit"])) - Decimal(str(m["debit"]))
        rows.append({
            "date": str(m["date"] or ""),
            "doc": m["doc"],
            "ref": m["ref"],
            "note": m["note"],
            "debit": _money(m["debit"]) if m["debit"] else "",
            "credit": _money(m["credit"]) if m["credit"] else "",
            "balance": _format_balance_with_label(balance),
        })

    if fmt == "pdf":
        return _statement_pdf(
            title="كشف حساب مورد",
            party_name=getattr(supplier, "name", str(supplier)),
            rows=rows,
            filename=f"supplier_statement_{supplier_id}.pdf",
            date_from=date_from_str,
            date_to=date_to_str,
        )

    return render(request, "accounting_app/supplier_statement.html", {
        "supplier": supplier,
        "rows": rows,
        "date_from": date_from_str,
        "date_to": date_to_str,
    })

from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.shortcuts import render, redirect
from .forms import AccountingPeriodForm

@login_required
@permission_required("accounting_app.add_accountingperiod", raise_exception=True)
def period_create(request):
    if request.method == "POST":
        form = AccountingPeriodForm(request.POST)
        if form.is_valid():
            p = form.save()
            messages.success(request, f"تم إنشاء الفترة المحاسبية: {p.name}")
            return redirect("account:periods_list")
    else:
        form = AccountingPeriodForm(initial={"is_closed": False})

    return render(request, "accounting_app/period_form.html", {"form": form})

