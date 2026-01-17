# inventory/utils_exports.py
import os
import io
import csv

from django.conf import settings
from django.http import HttpResponse
from django.utils import timezone

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

import arabic_reshaper
from bidi.algorithm import get_display

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# =========================
# ARABIC / PDF HELPERS
# =========================

def _ar(txt) -> str:
    return get_display(arabic_reshaper.reshape(str(txt or "")))


def _register_arabic_font() -> str:
    """
    Register Arabic font for ReportLab.
    Tries multiple known paths (including your old working one),
    and falls back to Windows Arial if missing.
    Returns the font name key to use in PDF ("AR_FONT").
    """
    # أهم مسار عندك كان شغّال (static/inventory/fonts)
    candidates = [
        os.path.join(settings.BASE_DIR, "inventory", "static", "inventory", "fonts", "Amiri-Regular.ttf"),
        os.path.join(settings.BASE_DIR, "inventory", "static", "fonts", "Amiri-Regular.ttf"),
        os.path.join(settings.BASE_DIR, "static", "fonts", "Amiri-Regular.ttf"),
    ]

    font_path = None
    for p in candidates:
        if os.path.exists(p):
            font_path = p
            break

    # fallback: Windows Arial (عشان ما يوقف التصدير)
    if not font_path:
        font_path = r"C:\Windows\Fonts\arial.ttf"

    # Register once
    try:
        pdfmetrics.getFont("AR_FONT")
    except Exception:
        pdfmetrics.registerFont(TTFont("AR_FONT", font_path))

    return "AR_FONT"


# =========================
# PDF EXPORT
# =========================

def build_products_pdf(rows, title="تقرير المنتجات"):
    """
    rows: list of dicts: {name, sku, unit, type_label}
    """
    font_name = _register_arabic_font()

    company_name = "شركة المحبة للصناعات الغذائية"
    company_address = "الأردن - عمّان - أبو علندا"
    now_str = timezone.localtime(timezone.now()).strftime("%Y-%m-%d %H:%M")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=24, rightMargin=24, topMargin=70, bottomMargin=45
    )

    styles = getSampleStyleSheet()
    arabic_style = ParagraphStyle(
        "arabic",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=12,
        leading=16,
        alignment=1  # center
    )
    arabic_title = ParagraphStyle(
        "arabic_title",
        parent=styles["Title"],
        fontName=font_name,
        fontSize=18,
        leading=22,
        alignment=1
    )
    arabic_small = ParagraphStyle(
        "arabic_small",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=10,
        leading=14,
        alignment=1
    )

    def header_footer(canvas, doc_):
        canvas.saveState()
        canvas.setFont(font_name, 12)
        canvas.drawCentredString(doc_.pagesize[0] / 2, doc_.pagesize[1] - 25, _ar(company_name))
        canvas.setFont(font_name, 10)
        canvas.drawCentredString(doc_.pagesize[0] / 2, doc_.pagesize[1] - 42, _ar(company_address))
        canvas.drawCentredString(doc_.pagesize[0] / 2, doc_.pagesize[1] - 58, _ar(f"تاريخ الطباعة: {now_str}"))
        canvas.setFont(font_name, 10)
        canvas.drawRightString(doc_.pagesize[0] - 24, 18, _ar(f"صفحة {doc_.page}"))
        canvas.restoreState()

    elements = []
    elements.append(Paragraph(_ar(title), arabic_title))
    elements.append(Paragraph(_ar(f"تاريخ ووقت الطباعة: {now_str}"), arabic_small))
    elements.append(Spacer(1, 12))

    headers = ["#", "اسم المنتج", "النوع", "SKU", "الوحدة"]
    data = [[Paragraph(_ar(h), arabic_style) for h in headers]]

    for i, r in enumerate(rows, start=1):
        data.append([
            Paragraph(_ar(i), arabic_style),
            Paragraph(_ar(r.get("name")), arabic_style),
            Paragraph(_ar(r.get("type_label")), arabic_style),
            Paragraph(_ar(r.get("sku")), arabic_style),
            Paragraph(_ar(r.get("unit")), arabic_style),
        ])

    # ✅ ===== تعديل تنسيق الجدول (عرض أقل + توسيط + نسب أعمدة) =====
    available_w = doc.width          # العرض داخل الهوامش
    table_w = available_w * 0.92     # خلي الجدول أضيق شوي (92%)

    # نسب الأعمدة (مجموعها = 1)
    col_ratios = [0.08, 0.38, 0.18, 0.22, 0.14]
    col_widths = [table_w * r for r in col_ratios]

    table = Table(
        data,
        colWidths=col_widths,
        repeatRows=1,
        hAlign="CENTER"  # توسيط الجدول
    )

    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 12),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E6E6E6")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
    ]))

    elements.append(table)

    doc.build(elements, onFirstPage=header_footer, onLaterPages=header_footer)
    buffer.seek(0)

    resp = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    resp["Content-Disposition"] = 'attachment; filename="products.pdf"'
    return resp


# =========================
# EXCEL HELPERS
# =========================

def style_sheet_like_pdf(ws, title, company_name, company_address, now_str):
    ws.sheet_view.rightToLeft = True

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill("solid", fgColor="E6E6E6")
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append([company_name])
    ws.append([company_address])
    ws.append([f"تاريخ ووقت الطباعة: {now_str}"])
    ws.append([title])
    ws.append([])

    for row in range(1, 5):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1).alignment = center
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)

    return bold, center, header_fill, border


def build_products_excel(wb, rows, title="تقرير المنتجات"):
    ws = wb.active
    ws.title = "المنتجات"

    company_name = "شركة المحبة للصناعات الغذائية"
    company_address = "الأردن - عمّان - أبو علندا"
    now_str = timezone.localtime(timezone.now()).strftime("%Y-%m-%d %H:%M")

    bold, center, header_fill, border = style_sheet_like_pdf(ws, title, company_name, company_address, now_str)

    headers = ["#", "اسم المنتج", "النوع", "SKU", "الوحدة"]
    ws.append(headers)

    header_row = ws.max_row
    for col in range(1, 6):
        c = ws.cell(row=header_row, column=col)
        c.font = bold
        c.alignment = center
        c.fill = header_fill
        c.border = border

    for i, r in enumerate(rows, start=1):
        ws.append([i, r.get("name"), r.get("type_label"), r.get("sku"), r.get("unit")])

    for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row, min_col=1, max_col=5):
        for c in row:
            c.alignment = center
            c.border = border

    widths = [6, 40, 18, 22, 12]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


# =========================
# CSV EXPORT
# =========================

def build_products_csv(products):
    """
    CSV = نص خام
    نخلّيه جدول نظيف يفتح مضبوط في Excel
    """
    response = HttpResponse(
        content_type="text/csv; charset=utf-8-sig"
    )
    response["Content-Disposition"] = 'attachment; filename="products.csv"'

    writer = csv.writer(response)

    # ✅ Header ثابت مثل PDF / Excel
    writer.writerow([
        "#",
        "اسم المنتج",
        "النوع",
        "SKU",
        "الوحدة",
    ])

    for i, p in enumerate(products, start=1):
        writer.writerow([
            i,
            getattr(p, "name", ""),
            getattr(p, "type_label", "") or getattr(p, "type", ""),
            getattr(p, "sku", "") or getattr(p, "code", ""),
            getattr(p, "unit", ""),
        ])

    return response
