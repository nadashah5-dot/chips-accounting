# inventory/utils.py
import io
import os
from datetime import datetime

import arabic_reshaper
from bidi.algorithm import get_display

from django.http import HttpResponse
from django.conf import settings

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont


def _ar(text: str) -> str:
    """شكل عربي + اتجاه صحيح"""
    return get_display(arabic_reshaper.reshape(str(text)))


class NumberedCanvasMixin:
    """
    Canvas helper لعرض: صفحة X من Y
    """
    def __init__(self, *args, **kwargs):
        from reportlab.pdfgen import canvas
        self._canvas = canvas.Canvas(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self._canvas.__dict__))
        self._canvas._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self._canvas.__dict__.update(state)
            self._draw_page_number(num_pages)
            self._canvas.showPage()
        self._canvas.save()

    def _draw_page_number(self, page_count):
        # ينعمل من callback لاحقاً
        pass

    def __getattr__(self, name):
        return getattr(self._canvas, name)


def export_warehouse_pdf_build(warehouse, rows):
    """
    rows = [{'product_name':..., 'quantity':..., 'code':...}, ...]
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=24,
        rightMargin=24,
        topMargin=60,
        bottomMargin=50,
    )

    # --- Register Arabic font ---
    font_path = os.path.join(settings.BASE_DIR, "inventory", "static", "inventory", "fonts", "Amiri-Regular.ttf")
    if not os.path.exists(font_path):
        # fallback to Windows Arial if font not found
        font_path = r"C:\Windows\Fonts\arial.ttf"

    pdfmetrics.registerFont(TTFont("AR_FONT", font_path))

    styles = getSampleStyleSheet()
    arabic_style = ParagraphStyle(
        "arabic",
        parent=styles["Normal"],
        fontName="AR_FONT",
        fontSize=12,
        leading=16,
        alignment=1,  # center
    )
    arabic_small = ParagraphStyle(
        "arabic_small",
        parent=styles["Normal"],
        fontName="AR_FONT",
        fontSize=10,
        leading=14,
        alignment=1,
    )

    company_name = "شركة المحبة للصناعات الغذائية"
    company_address = "الأردن - عمّان - أبو علندا"
    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    elements = []

    # عنوان
    elements.append(Paragraph(_ar(company_name), arabic_style))
    elements.append(Paragraph(_ar(company_address), arabic_small))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(_ar(f"تقرير مخزون المستودع: {warehouse.name}"), arabic_style))
    elements.append(Paragraph(_ar(f"تاريخ ووقت الطباعة: {now}"), arabic_small))
    elements.append(Spacer(1, 14))

    # جدول
    headers = [_ar("رقم المادة"), _ar("اسم المادة"), _ar("الكمية")]
    data = [[Paragraph(h, arabic_style) for h in headers]]

    total_qty = 0
    for r in rows:
        total_qty += int(r.get("quantity") or 0)
        data.append([
            Paragraph(_ar(r.get("code", "")), arabic_small),
            Paragraph(_ar(r.get("product_name", "")), arabic_small),
            Paragraph(_ar(r.get("quantity", "")), arabic_small),
        ])

    # سطر إجمالي
    data.append([
        Paragraph("", arabic_small),
        Paragraph(_ar("الإجمالي"), arabic_style),
        Paragraph(_ar(total_qty), arabic_style),
    ])

    table = Table(data, colWidths=[120, 380, 140], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2f3b46")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 1), (-1, -2), colors.whitesmoke),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#e9ecef")),
    ]))
    elements.append(table)

    # --- header/footer + page numbers ---
    def draw_header_footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("AR_FONT", 10)

        # Header line
        canvas.setStrokeColor(colors.grey)
        canvas.line(doc.leftMargin, doc.pagesize[1] - 45, doc.pagesize[0] - doc.rightMargin, doc.pagesize[1] - 45)

        # Footer line
        canvas.line(doc.leftMargin, 38, doc.pagesize[0] - doc.rightMargin, 38)

        # Footer text (company + address)
        canvas.drawCentredString(doc.pagesize[0] / 2, 22, _ar(f"{company_name} - {company_address}"))

        canvas.restoreState()

    class NumberedCanvas(NumberedCanvasMixin):
        def _draw_page_number(self, page_count):
            self.saveState()
            self.setFont("AR_FONT", 10)
            page_text = _ar(f"صفحة {self.getPageNumber()} من {page_count}")
            self.drawString(24, 44, page_text)  # فوق خط الفوتر
            self.restoreState()

    doc.build(
        elements,
        onFirstPage=draw_header_footer,
        onLaterPages=draw_header_footer,
        canvasmaker=NumberedCanvas,
    )

    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="warehouse_{warehouse.id}.pdf"'
    response.write(pdf)
    return response

# -------------------------------------------------
# receive_product
# -------------------------------------------------
from inventory.models import Warehouse, WarehouseStock, WarehouseMovement
from django.db import transaction

@transaction.atomic
def receive_product(product, qty, price=None, related_invoice=None):
    """
    تُستخدم عند إنشاء PurchaseItem
    تقوم بإضافة الكمية للمستودع الافتراضي وتسجيل حركة
    """

    # 1) اختيار مستودع افتراضي (أول مستودع)
    warehouse = Warehouse.objects.first()
    if not warehouse:
        return  # لو ما في مستودعات لا نعمل شي

    # 2) الحصول على الستوك أو إنشاؤه
    stock, _ = WarehouseStock.objects.get_or_create(
        warehouse=warehouse,
        product=product,
        defaults={'quantity': 0}
    )

    # 3) تحديث الكمية
    stock.quantity += qty
    stock.save()

    # 4) تسجيل حركة مستودع
    WarehouseMovement.objects.create(
        warehouse=warehouse,
        product=product,
        movement_type='إضافة',
        quantity=qty,
        notes=f"إضافة من فاتورة شراء {related_invoice or ''}"
    )

from io import BytesIO
from datetime import datetime

from django.http import HttpResponse

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

import arabic_reshaper
from bidi.algorithm import get_display


def _ar(txt: str) -> str:
    """Render Arabic correctly for ReportLab."""
    if txt is None:
        txt = ""
    reshaped = arabic_reshaper.reshape(str(txt))
    return get_display(reshaped)


def export_all_warehouses_pdf(warehouses_data: dict):
    """
    warehouses_data format:
    {
      "Warehouse Name": [{"product_name": "...", "quantity": 10, "code": "..."}, ...],
      ...
    }
    """
    buffer = io.BytesIO()

    # --- Register Arabic font (prefer Amiri) ---
    font_path = os.path.join(settings.BASE_DIR, "inventory", "static", "inventory", "fonts", "Amiri-Regular.ttf")
    if not os.path.exists(font_path):
        font_path = r"C:\Windows\Fonts\arial.ttf"

    pdfmetrics.registerFont(TTFont("AR_FONT", font_path))

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=24,
        rightMargin=24,
        topMargin=60,
        bottomMargin=50,
        title="All Warehouses Report"
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "title_ar",
        parent=styles["Normal"],
        fontName="AR_FONT",
        fontSize=16,
        leading=20,
        alignment=1,
    )
    normal_center = ParagraphStyle(
        "normal_center_ar",
        parent=styles["Normal"],
        fontName="AR_FONT",
        fontSize=11,
        leading=15,
        alignment=1,
    )
    normal_right = ParagraphStyle(
        "normal_right_ar",
        parent=styles["Normal"],
        fontName="AR_FONT",
        fontSize=11,
        leading=15,
        alignment=2,  # right
    )

    company_name = "شركة المحبة للصناعات الغذائية"
    company_address = "الأردن - عمّان - أبو علندا"
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    elements = []

    # محتوى التقرير: لكل مستودع جدول
    for w_name, rows in (warehouses_data or {}).items():

        elements.append(Paragraph(_ar(company_name), title_style))
        elements.append(Paragraph(_ar(company_address), normal_center))
        elements.append(Spacer(1, 8))

        elements.append(Paragraph(_ar(f"تقرير مخزون المستودع: {w_name}"), title_style))
        elements.append(Paragraph(_ar(f"تاريخ ووقت الطباعة: {now_str}"), normal_center))
        elements.append(Spacer(1, 14))

        headers = [_ar("رمز المنتج"), _ar("اسم المنتج"), _ar("الكمية")]
        data = [[Paragraph(h, normal_center) for h in headers]]

        total_qty = 0
        for r in rows:
            qty = r.get("quantity", 0) or 0
            try:
                qty_int = int(qty)
            except Exception:
                qty_int = 0
            total_qty += qty_int

            data.append([
                Paragraph(_ar(r.get("code", "")), normal_center),
                Paragraph(_ar(r.get("product_name", "")), normal_right),
                Paragraph(_ar(str(qty)), normal_center),
            ])

        data.append([
            Paragraph("", normal_center),
            Paragraph(_ar("الإجمالي"), normal_center),
            Paragraph(_ar(str(total_qty)), normal_center),
        ])

        table = Table(data, colWidths=[140, 420, 140], repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2f3b46")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("BACKGROUND", (0, 1), (-1, -2), colors.whitesmoke),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#e9ecef")),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 16))

        # صفحة جديدة بين المستودعات
        from reportlab.platypus import PageBreak
        elements.append(PageBreak())

    # Header/Footer + Page X of Y
    def draw_header_footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("AR_FONT", 10)

        # خطوط
        canvas.setStrokeColor(colors.grey)
        canvas.line(doc.leftMargin, doc.pagesize[1] - 45, doc.pagesize[0] - doc.rightMargin, doc.pagesize[1] - 45)
        canvas.line(doc.leftMargin, 38, doc.pagesize[0] - doc.rightMargin, 38)

        # فوتر ثابت
        canvas.drawCentredString(doc.pagesize[0] / 2, 22, _ar(f"{company_name} - {company_address}"))
        canvas.restoreState()

    class NumberedCanvas(NumberedCanvasMixin):
        def _draw_page_number(self, page_count):
            self.saveState()
            self.setFont("AR_FONT", 10)
            page_text = _ar(f"صفحة {self.getPageNumber()} من {page_count}")
            self.drawString(24, 44, page_text)
            self.restoreState()

    doc.build(
        elements,
        onFirstPage=draw_header_footer,
        onLaterPages=draw_header_footer,
        canvasmaker=NumberedCanvas,
    )

    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="all_warehouses.pdf"'
    response.write(pdf)
    return response

from datetime import datetime
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.worksheet.page import PageMargins

def build_warehouse_excel(warehouse_name: str, rows: list, filename: str):
    """
    rows: [{"product_name": "...", "quantity": 10, "code": "..."}]
    """
    company_name = "شركة المحبة للصناعات الغذائية"
    company_address = "الأردن - عمّان - أبو علندا"
    now_str = timezone.localtime(timezone.now()).strftime("%Y-%m-%d %H:%M:%S")

    wb = Workbook()
    ws = wb.active
    ws.title = "تقرير المستودع"
    ws.sheet_view.rightToLeft = True

    # Print settings
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.6, bottom=0.6)

    # Header / Footer for printing
    ws.oddHeader.center.text = company_name
    ws.oddHeader.right.text = f"تاريخ الطباعة: {now_str}"
    ws.oddFooter.center.text = f"{company_name} - {company_address}"
    ws.oddFooter.right.text = "صفحة &P من &N"

    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="2F3B46")
    alt_fill = PatternFill("solid", fgColor="F5F5F5")

    title_font = Font(bold=True, size=16)
    subtitle_font = Font(bold=True, size=11)
    header_font = Font(bold=True, color="FFFFFF")

    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    # Top header area (merged)
    ws.merge_cells("A1:D1")
    ws["A1"] = company_name
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    ws.merge_cells("A2:D2")
    ws["A2"] = company_address
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = center

    ws.merge_cells("A3:D3")
    ws["A3"] = f"المستودع: {warehouse_name}"
    ws["A3"].font = subtitle_font
    ws["A3"].alignment = center

    ws.merge_cells("A4:D4")
    ws["A4"] = f"تاريخ ووقت الطباعة: {now_str}"
    ws["A4"].alignment = center

    start_row = 6

    # Table header
    headers = ["#", "رمز المنتج", "اسم المنتج", "الكمية"]
    for col, val in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=val)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # Data
    total_qty = 0
    for i, r in enumerate(rows, 1):
        row_idx = start_row + i
        code = r.get("code", "") or ""
        name = r.get("product_name", "") or ""
        qty = int(r.get("quantity") or 0)
        total_qty += qty

        data = [i, code, name, qty]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = right if col == 3 else center
            cell.border = border
            if i % 2 == 0:
                cell.fill = alt_fill

    # Total row
    total_row = start_row + len(rows) + 1
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
    ws.cell(row=total_row, column=1, value="الإجمالي").alignment = center
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    ws.cell(row=total_row, column=1).border = border

    ws.cell(row=total_row, column=4, value=total_qty).alignment = center
    ws.cell(row=total_row, column=4).font = Font(bold=True)
    ws.cell(row=total_row, column=4).border = border

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 14

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

    ...
